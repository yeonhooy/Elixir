from __future__ import absolute_import, division, print_function, unicode_literals, unicode_literals

import pathlib
from collections import Counter

import matplotlib.pyplot as plt

import numpy as np
import pandas as pd
import tensorflow as tf
from tensorflow import keras
from tensorflow.keras.regularizers import l1, l2
from tensorflow.keras.optimizers import SGD, Adam, RMSprop
from tensorflow.keras.layers import Dropout
from tensorflow.keras.models import Sequential
from tensorflow.keras import layers
from tensorflow.keras.layers import Dense
from tensorflow.keras.layers import Flatten
from tensorflow.keras.layers import Conv1D
from tensorflow.keras.layers import MaxPooling1D

from sklearn.pipeline import make_pipeline, Pipeline
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
from sklearn.svm import LinearSVR
from sklearn.multioutput import MultiOutputRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import mean_absolute_error, accuracy_score, mean_squared_error
from sklearn.preprocessing import StandardScaler, MinMaxScaler, RobustScaler
from sklearn.model_selection import KFold
from sklearn.metrics import r2_score

import lightgbm as lgb

import xgboost as xgb

import random
import sys
from itertools import combinations

from scipy import stats

# Ignore all GPUs, tf random forest does not benefit from it.
import os
import shutil
import datetime
# os.environ["TF_CPP_MIN_LOG_LEVEL"] = "2"
# os.environ["CUDA_VISIBLE_DEVICES"] = ""

from openpyxl import Workbook, load_workbook


class PrintDot(keras.callbacks.Callback):
    def on_epoch_end(self, epoch, logs):
        if epoch % 100 == 0: print('')
        print('~', end='')  # print ~ when 1 epoch end


# 1.functions
# <1. detect outlier and remove outliers>
def detect_outliers(df, n, features):
    outlier_indices = []
    print("df", df)
    for col in features:
        print("col", col)  # [Output]
        Q1 = np.percentile(df[col], 25)
        Q3 = np.percentile(df[col], 75)
        print("Q1", Q1)
        print("Q3", Q3)
        IQR = Q3 - Q1
        outlier_step = 1.5 * IQR
        minQ = Q1 - outlier_step
        maxQ = Q3 + outlier_step
        print("minQ", minQ)
        print("maxQ", maxQ)
        outlier_list_col = df[(df[col] < minQ) | (df[col] > maxQ)].index
        outlier_indices.extend(outlier_list_col)
        print("Outlier`s len", len(outlier_indices))
        print("Outlier", outlier_indices)
    # outlier_indices = Counter(outlier_indices)
    # print("outlier_indices.items()",outlier_indices)
    # multiple_outliers = list(k for k, v in outlier_indices.items() if v > n)

    return outlier_indices


def detect_outliers_delete(df, n, features):
    outlier_indices = []
    print("df", df)
    for col in features:
        print("col", col)  # [Output]
        Q1 = np.percentile(df[col], 25)
        Q3 = np.percentile(df[col], 75)
        print("Q1", Q1)
        print("Q3", Q3)
        IQR = Q3 - Q1
        outlier_step = 1.5 * IQR
        minQ = Q1 - outlier_step
        maxQ = Q3 + outlier_step
        print("minQ", minQ)
        print("maxQ", maxQ)
        outlier_list_col = df[(df[col] < minQ) | (df[col] > maxQ)].index
        outlier_indices.extend(outlier_list_col)
        print("Outlier`s len", len(outlier_indices))
        print("Outlier", outlier_indices)
        for i in range(len(outlier_indices)):
            df = df.drop(df.index[outlier_indices[i]])
        print("new df ", df)
    # outlier_indices = Counter(outlier_indices)
    # print("outlier_indices.items()",outlier_indices)
    # multiple_outliers = list(k for k, v in outlier_indices.items() if v > n)

    return df


def print_accuracy_all_single(test_predictions, y_test):
    # 출력결과 배열로 저장
    result = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    # Accuracy 측정
    dot_70 = 0
    dot_80 = 0
    dot_90 = 0
    abserror = np.abs((test_predictions - y_test.array))
    errorrate = abserror / (y_test.array) * 100
    maperate = np.abs((test_predictions - y_test.array) / y_test.array) * 100
    accrate = 100 - errorrate

    for u in accrate:
        if (u >= 70 and u <= 100):
            dot_70 = dot_70 + 1
    result[0] = round(dot_70 / len(accrate) * 100, 3)
    print("OVER 70: ")
    print(result[0])

    for q in accrate:
        if (q >= 80 and q <= 100):
            dot_80 = dot_80 + 1
    result[1] = round(dot_80 / len(accrate) * 100, 3)
    print("OVER 80: ")
    print(result[1])

    for z in accrate:
        if (z >= 90 and z <= 100):
            dot_90 = dot_90 + 1
    result[2] = round(dot_90 / len(accrate) * 100, 3)
    print("OVER 90: ")
    print(result[2])

    # MSE
    n_mse = len(accrate)
    square_mse = np.square(abserror)
    result[3] = round(np.sum(square_mse) / n_mse, 3)
    print("MSE: ")
    print(result[3])

    # RMSE
    n_rmse = len(accrate)
    mse = np.square(abserror)
    smse = np.sum(mse) / n_rmse
    result[4] = round(np.sqrt(smse), 3)
    print("RMSE: ")
    print(result[4])

    # MAE
    n_mae = len(accrate)
    result[5] = round(np.sum(abserror) / n_mae, 3)
    print("Mae: ")
    print(result[5])

    # MAPE
    n_mape = len(accrate)
    result[6] = round(np.sum(maperate) / n_mape, 3)
    print("MAPE: ")
    print(result[6])

    # Pearson R
    # val_sum = np.sum(test_predictions)
    # array_has_stranges = np.isnan(val_sum) + np.isinf(val_sum)
    # val_sum2 = np.sum(y_test)
    # arrayhas_stranges2 = np.isnan(val_sum2) + np.isinf(val_sum2)

    if any(np.isnan(val) for val in test_predictions.astype(np.float32)) or any(
            np.isnan(val) for val in y_test.astype(np.float32)) or any(
            np.isinf(val) for val in test_predictions.astype(np.float32)) or any(
            np.isinf(val) for val in y_test.astype(np.float32)):
        # test_predictions.fillna(test_predictions.mean())
        # y_test.fillna(y_test.mean())
        result[7] = 0
    else:
        pearsonr = stats.pearsonr(y_test, test_predictions)
        result[7] = round(pearsonr[0], 3)

    print("Pearson R: ")
    print(result[7])

    # R2

    if any(np.isnan(val) for val in test_predictions.astype(np.float32)) or any(
            np.isnan(val) for val in y_test.astype(np.float32)) or any(
            np.isinf(val) for val in test_predictions.astype(np.float32)) or any(
            np.isinf(val) for val in y_test.astype(np.float32)):
        # test_predictions.fillna(test_predictions.mean())
        # y_test.fillna(y_test.mean())
        result[8] = 0
    else:
        r2cal = r2_score(y_test, test_predictions)
        result[8] = round(r2cal, 3)

    print("R2: ")
    print(result[8])

    return result


# <2. print result with 5 metrics>
def print_accuracy_all(test_predictions, y_test, modelName):
    # todo: 정확도 결과들을 딕셔너리 형태로 변경하기 (output마다 result를 한번에 저장하기 위함)

    # 출력결과 배열로 저장
    # todo : result 데이터형태를 딕셔너리 형태로
    result_1 = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    result_2 = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    result_3 = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    result_4 = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    # result = [[0, 0, 0, 0, 0, 0, 0, 0, 0],[0, 0, 0, 0, 0, 0, 0, 0, 0],
    #           [0, 0, 0, 0, 0, 0, 0, 0, 0],[0, 0, 0, 0, 0, 0, 0, 0, 0]]

    result = {
        'Recv_bw': result_1,
        'Sent_bw': result_2,
        'Max_sent': result_3,
        'Max_recv': result_4
    }

    # Accuracy 측정
    dot_70 = 0
    dot_80 = 0
    dot_90 = 0
    # print("_____________________________________")
    # print("to check test / predict matirx type")
    # print(test_predictions)
    # print(y_test)
    # print("_____________________________________")
    abserror = np.abs((test_predictions - y_test))
    errorrate = abserror / (y_test) * 100
    maperate = np.abs((test_predictions - y_test) / y_test) * 100
    accrate = 100 - errorrate
    print("accrate!!!")
    acc_array = np.array(accrate)

    print(test_predictions)
    print(y_test)

    predict_col_1 = np.array(test_predictions).T[0]
    predict_col_2 = np.array(test_predictions).T[1]
    predict_col_3 = np.array(test_predictions).T[2]
    predict_col_4 = np.array(test_predictions).T[3]

    # print(predict_col_1,predict_col_2,predict_col_3,predict_col_4)

    test_col_1 = y_test.iloc[:, 0]
    test_col_2 = y_test.iloc[:, 1]
    test_col_3 = y_test.iloc[:, 2]
    test_col_4 = y_test.iloc[:, 3]

    test_col_1 = np.array(test_col_1).flatten()
    test_col_2 = np.array(test_col_2).flatten()
    test_col_3 = np.array(test_col_3).flatten()
    test_col_4 = np.array(test_col_4).flatten()

    acc_col_1 = accrate.iloc[:, 0]
    acc_col_2 = accrate.iloc[:, 1]
    acc_col_3 = accrate.iloc[:, 2]
    acc_col_4 = accrate.iloc[:, 3]

    acc_col_1 = np.array(acc_col_1).flatten()
    acc_col_2 = np.array(acc_col_2).flatten()
    acc_col_3 = np.array(acc_col_3).flatten()
    acc_col_4 = np.array(acc_col_4).flatten()

    print(modelName)
    result = [calculate_accrate('Sent_bw', predict_col_1, test_col_1, acc_col_1, result_1),
              calculate_accrate('Recv_bw', predict_col_2, test_col_2, acc_col_2, result_2),
              calculate_accrate('Max_sent', predict_col_3, test_col_3, acc_col_3, result_3),
              calculate_accrate('Max_recv', predict_col_4, test_col_4, acc_col_4, result_4)]

    print(result)
    return result


def calculate_accrate(name, predict, test, accrate, result):
    # print(name)
    dot_70 = 0
    dot_80 = 0
    dot_90 = 0
    for u in accrate:
        if (u >= 70 and u <= 100):
            dot_70 = dot_70 + 1
    result[0] = round(dot_70 / len(accrate) * 100, 3)
    # print("OVER 70: ")
    # print(result[0])

    for q in accrate:
        if (q >= 80 and q <= 100):
            dot_80 = dot_80 + 1
    result[1] = round(dot_80 / len(accrate) * 100, 3)
    # print("OVER 80: ")
    # print(result[1])

    for z in accrate:
        if (z >= 90 and z <= 100):
            dot_90 = dot_90 + 1
    result[2] = round(dot_90 / len(accrate) * 100, 3)
    # print("OVER 90: ")
    # print(result[2])

    abserror = np.abs((predict - test))
    maperate = np.abs((predict - test) / test) * 100

    # MSE
    n_mse = len(test)
    square_mse = np.square(abserror)
    result[3] = round(np.sum(square_mse) / n_mse, 3)
    # print("MSE: ")
    # print(result[3])

    # RMSE
    n_rmse = len(test)
    mse = np.square(abserror)
    smse = np.sum(mse) / n_rmse
    result[4] = round(np.sqrt(smse), 3)
    # print("RMSE: ")
    # print(result[4])

    # MAE
    n_mae = len(test)
    result[5] = round(np.sum(abserror) / n_mae, 3)
    # print("Mae: ")
    # print(result[5])

    # MAPE
    n_mape = len(test)
    result[6] = round(np.sum(maperate) / n_mape, 3)
    # print("MAPE: ")
    # print(result[6])

    # Pearson R
    if any(np.isnan(val) for val in predict.astype(np.float32)) or any(
            np.isnan(val) for val in test.astype(np.float32)) or any(
        np.isinf(val) for val in predict.astype(np.float32)) or any(
        np.isinf(val) for val in test.astype(np.float32)):
        # test_predictions.fillna(test_predictions.mean())
        # y_test.fillna(y_test.mean())
        # result[7] = 'NA'
        result[7] = 0
    else:
        pearsonr = stats.pearsonr(test, predict)
        result[7] = round(pearsonr[0], 3)

    # print("Pearson R: ")
    # print(result[7])

    # R2
    if any(np.isnan(val) for val in predict.astype(np.float32)) or any(
            np.isnan(val) for val in test.astype(np.float32)) or any(
        np.isinf(val) for val in predict.astype(np.float32)) or any(
        np.isinf(val) for val in test.astype(np.float32)):
        # test_predictions.fillna(test_predictions.mean())
        # y_test.fillna(y_test.mean())
        # result[8] = 'NA'
        result[8] = 0

    else:
        r2cal = r2_score(test, predict)
        result[8] = round(r2cal, 3)

    # print("R2: ")
    # print(result[8])

    return result


# <3. show predict accuracy graph>
def plot_history(history):
    hist = pd.DataFrame(history.history)
    hist['epoch'] = history.epoch

    plt.figure(figsize=(6, 9))

    plt.subplot(2, 1, 1)
    plt.xlabel('Epoch')
    plt.ylabel('Mean Abs Error [Output]')
    plt.plot(hist['epoch'], hist['mean_absolute_error'],
             label='Train Error')
    plt.plot(hist['epoch'], hist['val_mean_absolute_error'],
             label='Val Error')
    plt.ylim([0, 5])
    plt.legend()

    plt.subplot(2, 1, 2)
    plt.xlabel('Epoch')
    plt.ylabel('Mean Square Error [Output^2$]')
    plt.plot(hist['epoch'], hist['mean_squared_error'],
             label='Train Error')
    plt.plot(hist['epoch'], hist['val_mean_squared_error'],
             label='Val Error')
    plt.ylim([0, 20])
    plt.legend()
    plt.show()


# 4. Write result to excel file

def write_excel(result_excel, models, results, result_name, outputTitles):
    # outputTitles = ['Sent_bw','Recv_bw', 'Max_sent', 'Max_recv']
    print("NOW SAVE ", result_name, results)
    print("TEST: ", results[0][0][0])
    for o in range(0, 4):
        eIndex = o + 1
        wTitle = outputTitles[o]
        result_excel.create_sheet(index=eIndex, title=wTitle)
        ws_1 = result_excel[wTitle]  # excel 파일에 ouptut label 별로 worksheet 따로 작성되는것 같음
        ws_1.cell(row=1, column=1).value = 'Model'
        ws_1.cell(row=1, column=2).value = 'Layer'  # 1행, 2열부터 13열까지 label을 작성
        ws_1.cell(row=1, column=3).value = 'Nodes'
        ws_1.cell(row=1, column=4).value = 'Over 70'
        ws_1.cell(row=1, column=5).value = 'Over 80'
        ws_1.cell(row=1, column=6).value = 'Over 90'
        ws_1.cell(row=1, column=7).value = 'MSE'
        ws_1.cell(row=1, column=8).value = 'RMSE'
        ws_1.cell(row=1, column=9).value = 'MAE'
        ws_1.cell(row=1, column=10).value = 'MAPE'
        ws_1.cell(row=1, column=11).value = 'Pearson R'
        ws_1.cell(row=1, column=12).value = 'Pearson R2'
        ws_1.cell(row=1, column=13).value = 'Design'
        counter = 2

        model_len = len(models)
        print("LENMODEL: ", model_len)
        for i in range(len(models)):  # model의 길이만큼 밑으로 작성
            print("MODEL_SUMMARY: ", str(models[i][4]))
            ws_1.cell(row=counter, column=1).value = models[i][0]  # model 의 이름
            ws_1.cell(row=counter, column=2).value = models[i][1]  # model의 layer 수
            # ws_1.cell(row=counter, column=3).value = models[i][2]  #model의 nodes 수
            ws_1.cell(row=counter, column=3).value = str(models[i][3])  # model의 design / type = str
            for j in range(0, 9):
                colj = 4 + j
                print(i, o, j)
                ws_1.cell(row=counter, column=colj).value = results[i][o][j]  # result 총 9개
            ws_1.cell(row=counter, column=13).value = str(models[i][4])  # model의 design / type = str
            counter = counter + 1


# 2. Models


# 1) <Linear Regresssion>  - /multiple output regression /
def lrModel(X_train, X_test, y_train, y_test):
    # 스케일링
    # 1.minmax 스케일러
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)

    print(X_train, len(X_train))
    print(X_test, len(X_test))

    # 2. Build linearRegresion
    model = Sequential()
    model.add(Dense(4, activation='linear', input_dim=X_train.shape[1]))
    # model.add(Dense(4))

    optimizer = tf.train.GradientDescentOptimizer(learning_rate=0.001)
    # optimizer = tf.keras.optimizers.RMSprop(0.001)

    model.compile(loss='mean_squared_error',
                  optimizer=optimizer,
                  metrics=['mean_absolute_error', 'mean_squared_error'])

    print("<<<<<LR>>>>>")
    model.summary()
    # patience 매개변수는 성능 향상을 체크할 에포크 횟수입니다
    early_stop = keras.callbacks.EarlyStopping(monitor='val_loss', patience=10)

    model.fit(
        X_train, y_train.to_numpy(),
        epochs=1000, batch_size=10, validation_split=0.2, verbose=0,
        callbacks=[early_stop, PrintDot()])

    # loss, mean_absolute_error, mean_squared_error = lRmodel.evaluate(X_test, y_test.to_numpy(), verbose=2)

    test_predictions = model.predict(X_test)
    print(X_test)
    print(
        "***********************************************************************************************************************************",
        test_predictions)
    result = print_accuracy_all(test_predictions, y_test, 'Linear Regression')

    # Save Trained Model
    # model.save('alpha_LR.h5')

    return result


def linearRegression_sk(inferencefile, X_train, X_test, y_train, y_test, test_input, outputType):
    # 스케일링
    # 1.minmax 스케일러
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    ptest_input = scaler.transform(test_input)

    model = make_pipeline(RobustScaler(), MultiOutputRegressor(LinearRegression()))
    model.fit(X_train, y_train)

    test_predictions = model.predict(X_test)
    inference = model.predict(ptest_input)

    # wirteInference(test_input,inference,inferencefile,outputType,"LR")

    return inference


# 2) <Linear SVR>  / single output - todo : sinlge --> mutliple outptut / (done)
def lsvrModel(inferencefile, X_train, X_test, y_train, y_test, test_input, outputType):
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    ptest_input = scaler.transform(test_input)
    # test_input = scaler.transform(test_input)
    model = make_pipeline(RobustScaler(), MultiOutputRegressor(LinearSVR(random_state=0, tol=1e-05, max_iter=5000)))
    # model.fit(X_train, y_train.values.ravel())
    model.fit(X_train, y_train)
    test_predictions = model.predict(X_test)
    # result = print_accuracy_all(test_predictions, y_test,"Linear SVR")
    inference = model.predict(ptest_input)

    # wirteInference(test_input, inference, inferencefile, outputType, "SVR")

    return inference


# 3) <ForestRegression>
def rforestModel(inferencefile, X_train, X_test, y_train, y_test, tree, test_input, outputType):
    # 스케일링
    # 1.minmax 스케일러
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    ptest_input = scaler.transform(test_input)

    # 2. build model
    model = RandomForestRegressor(n_estimators=tree)
    model.fit(X_train, y_train)

    # 3.Predict & Test the Model ############
    real_output = np.array(X_test)
    inputs = real_output
    predicted_output_3 = model.predict(inputs)
    # test_predictions_3 = predicted_output_3.flatten()
    # result = print_accuracy_all(predicted_output_3, y_test,"Random Forest")

    # inference
    test_output = np.array(ptest_input)
    inference = model.predict(test_output)

    # wirteInference(test_input, inference, inferencefile, outputType, "RF")

    # Save Trained Model
    # model.save('alpha_RF.h5')

    return inference


def lightgbm_model(X_train, X_test, y_train, y_test, test_input):
    print("*******lightgbm**********")
    X_train, X_valid, y_train, y_valid = train_test_split(X_train, y_train, test_size=0.1, random_state=20)
    train_data = lgb.Dataset(X_train, label=y_train)
    valid_data = lgb.Dataset(X_valid, label=y_valid)
    inf_dta = lgb.Dataset(test_input)
    print(y_valid)

    params = {'learning_rate': 0.01,
              'max_depth': 16,
              'boosting': 'gbdt',
              'objective': 'regression',
              'metric': ['mse', 'mae', 'mape'],
              'is_training_metric': True,
              'num_leaves': 144,
              'feature_fraction': 0.9,
              'bagging_fraction': 0.7,
              'bagging_freq': 5,
              'seed': 2018}

    # num_round = 10

    model = lgb.train(params, train_data, 1000, valid_sets=[valid_data], early_stopping_rounds=100)

    # predict_test = model.predict(X_test)

    inference = model.predict(test_input)

    return inference


def XGBoost_model(X_train, X_test, y_train, y_test, test_input):
    print("*******XGBOOST MODEL*******")
    train_data = xgb.DMatrix(data=X_train, label=y_train)
    test_data = xgb.DMatrix(data=X_test, label=y_test)
    inf_dta = xgb.DMatrix(data=test_input)
    param = {'max_depth': 8, 'objective': 'reg:squarederror', 'eval_metric': ['rmse', 'mae']}
    num_round = 10

    # BASIC MODEL
    bst = xgb.train(param, train_data, num_round)
    # make prediction
    # preds = bst.predict(test_data)

    inference = bst.predict(inf_dta)

    # RF MODEL
    #    bst = xgb.XGBRFRegressor(random_state=42).fit(X_train, y_train)
    #    preds = bst.predict(X_test)

    return inference


# 4) <ANN>
class Architecture:  # todo: architecture 에 다른 parameter 추가 (10.21)
    def __init__(self, name, layer, nodes, design):
        self.name = name
        self.layer = layer
        self.nodes = nodes
        self.units = []
        self.modelSummary = []
        self.design = design

    def initial_utnis(self, units):
        self.units = units

    def update_name(self, name):
        self.name = name
        # self.units = self.design.nodesUnit(self.layer, self.nodes)

    def update_nodes(self, nodes):
        self.nodes = nodes
        self.units = self.design.nodesUnit(self.layer, self.nodes, self.units)

    def update_units(self, design_type):
        self.design.type = design_type
        self.units = self.design.nodesUnit(self.layer, self.nodes, self.units)

    def update_layers(self, layer):
        self.layer = layer
        self.units = self.design.nodesUnit(layer, 0, self.units)

    def update_modelSumary(self, modelSummary):
        self.modelSummary = modelSummary

    def get_arch(self):
        tmp = [self.name, self.layer, self.nodes, self.units, self.modelSummary]
        # tmp = tmp + self.design.get_design()
        return tmp


class Design:
    def __init__(self):
        self.type
        self.dropout
        self.regularizer
        self.initializer
        self.lr
        self.beta_1
        self.beta_2
        self.epsilon
        self.decay
        self.optimizer

    def nodesUnit(self, layer, nodes, units):
        if self.type == 'Default':
            print("info-", layer, nodes, units)
            if nodes != 0:
                units.append(nodes)
            print("(Defalut)First line Tmp : ", units)
            # tmp.append(4)  #for output node?
            print("(Defalut) Tmp : ", units)
            return units
        if self.type == 'First_half':
            tmp = [i for i in range(4, nodes + 1)]
            tmp = tmp + [nodes for i in range(layer - nodes - 1)]
            print("(FirstHalf)First line Tmp : ", tmp)
            # tmp.append(4)  # for output node?
            print("(FirstHalf)Second line Tmp : ", tmp)
            return tmp

    def get_design(self):
        return [self.type, self.dropout, self.regularizer,
                self.initializer, self.lr, self.beta_1,
                self.beta_2, self.epsilon, self.decay, self.optimizer]


class Default(Design):
    def __init__(self):
        self.name = 'ANN'
        self.type = 'Default'
        self.dropout = 0.2
        self.regularizer = l2(0.001)
        self.initializer = ['he_normal', 'he_uniform']
        self.activations = ['elu', 'relu', 'linear']
        self.lossfunctions = ['msle', 'mse', 'mae', 'mape']
        self.lr = 0.001
        self.beta_1 = 0.9
        self.beta_2 = 0.999
        self.epsilon = 1e-08
        self.decay = 0.0
        self.optimizer = Adam(lr=self.lr, beta_1=self.beta_1, beta_2=self.beta_2, epsilon=self.epsilon,
                              decay=self.decay)


def annModel(X_train, X_test, y_train, y_test, test_input, annname, outputType, resultFileName):
    # 스케일링
    # 1.minmax 스케일러
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    test_input = scaler.transform(test_input)
    modelSummary = []

    loadmodel_path = "dataset_new_avg/elixir-modelsave/%s_%s_%s.h5" % (resultFileName, outputType, annname)
    try:
        model = keras.models.load_model(loadmodel_path)
    except:
        print("no model!!")
        return None, None, None

    print("<<<<<<ANN_space>>>>>>>>>>>")
    model.summary()

    ############ Predict & Test the Model ############
    # real_output = np.array(X_test)
    # inputs = real_output

    # input("input")

    # inference
    test_output = np.array(test_input)
    inference = model.predict(test_output)

    #input(len(inference))
    # print("inference", inference)

    # print accuracy
    # result = print_accuracy_all(predicted_output, y_test,"ANN")
    return inference

def cnnModel(X_train, X_test, y_train, y_test, test_input, cnnname, outputType, resultFileName):
    # 스케일링
    # 1.minmax 스케일러
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    test_input = scaler.transform(test_input)
    modelSummary = []

    X_train_flat, X_test_flat, test_input, convnode = convolutionLayer(X_train,X_test,test_input,cnnname,outputType,resultFileName)
    print("convnode: ",convnode)

    loadmodel_path = "dataset_new_avg/elixir-modelsave/%s_%s_%s.h5" % (resultFileName, outputType, cnnname)
    try:
        model = keras.models.load_model(loadmodel_path)
    except:
        print("no model!!")
        return None, None, None

    print("<<<<<<ANN_space>>>>>>>>>>>")
    model.summary()

    ############ Predict & Test the Model ############
    # real_output = np.array(X_test)
    # inputs = real_output

    # input("input")

    # inference
    test_output = np.array(test_input)
    inference = model.predict(test_output)

    #input(len(inference))
    # print("inference", inference)

    # print accuracy
    # result = print_accuracy_all(predicted_output, y_test,"ANN")
    return inference

def convolutionLayer(X_train,X_test,X_valid,cnnname,outputType,resultFileName):
    # scale
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.fit_transform(X_test)
    X_valid = scaler.fit_transform(X_valid)
    X_train_conv = []
    X_test_conv = []
    X_valid_conv = []
    X_train_flat = []
    X_test_flat = []
    X_valid_flat = []
    # conv layer
    # filter size
    # filter Num
    # stride size
    # pooling layer (max or average)
    convparams = []  # [filtersize, filterNum, strideSize, pooling layer)
    print("x_train.shape")
    print(X_train.shape)
    print(X_train.shape[0])

    convSummary = []
    intKernel = 7
    nextKernel = 7
    intFilter = 1
    kernelShape = (1, intKernel, intFilter)
    filterNum = 1

    #1. data convert reshape: 3d form (1,7,1)
    for q in range(0, X_train.shape[0]):
        print(X_train[q].shape)
        print(q)
        X_train_conv.append(X_train[q].reshape(1, 7, 1))
        print(X_train[q].reshape(1, 7, 1))
    for k in range(0, X_test.shape[0]):
        X_test_conv.append(X_test[k].reshape(1, 7, 1))
    for v in range(0, X_valid.shape[0]):
        X_valid_conv.append(X_valid[v].reshape(1, 7, 1))

    #2. read conv summary
    while(True):
        convsumPath = "dataset_new_avg/elixir-modelsave/%s_%s_cnn_convsummary_elixir.txt" % (resultFileName, outputType)
        convsumtxt = open(convsumPath,'r')
        sumlines = []
        thisconvSum = []
        findname = 0
        for paragraph in convsumtxt:
            sumlines = paragraph.split("/")
            for each_line in sumlines:
                if each_line==cnnname:
                    thisconvSum=sumlines
                    findname = findname+1
                    break
                else:
                    pass
        convsumtxt.close()
        print(resultFileName,outputType)
        print(cnnname)
        if findname > 0:
            break
    print(thisconvSum)

    layersummary = thisconvSum[2]
    cx = "[]()"
    arraysummary = ''.join(x for x in layersummary if x not in cx)
    print(arraysummary)
    summaries = arraysummary.split(", ")
    print(summaries)
    print(len(summaries))
    layertupples = []
    tuplen = int(len(summaries)/7)
    for q in range(tuplen):
        tup = (summaries[0+7*q],summaries[1+7*q],summaries[2+7*q],summaries[3+7*q],summaries[4+7*q],summaries[5+7*q],summaries[6+7*q])
        layertupples.append(tup)
    print(layertupples)

    kernelShape = (1,intKernel, intFilter)
    if tuplen >1:
        for tu in range(tuplen-1):
            filterNum = int(layertupples[tu][0])
            filter = int(layertupples[tu][1])
            acf = str(layertupples[tu][2])
            stride = int(layertupples[tu][3])
            flatcheck = str(layertupples[tu][4])
            intKernel = int(layertupples[tu][5])
            nextKernel = int(layertupples[tu][6])
            outShpae = (1, nextKernel, filterNum)
            print(acf)
            print(type(acf))

            for j in range(0, X_train.shape[0]):
                X_train_conv[j] = Conv1D(filters=filterNum,kernel_size=filter,strides=stride,input_shape=kernelShape)(X_train_conv[j])
            for m in range(0, X_test.shape[0]):
                X_test_conv[m] = Conv1D(filters=filterNum, kernel_size=filter, strides=stride, input_shape=kernelShape)(X_test_conv[m])
            for v in range(0, X_valid.shape[0]):
                X_valid_conv[v] = Conv1D(filters=filterNum, kernel_size=filter, strides=stride, input_shape=kernelShape)(X_valid_conv[v])
            print(X_train_conv[0])
            kernelShape = outShpae

    lastNodes = nextKernel * filterNum
    if str(type(X_train_conv[0])) == "<class 'numpy.ndarray'>":
        for ia in range(0, X_train.shape[0]):
            X_train_flat.append(X_train_conv[ia].flatten())
            #X_train_flat.append(X_train_conv[ia].reshape(1,lastNodes))
        for k in range(0, X_test.shape[0]):
            X_test_flat.append(X_test_conv[k].flatten())
            #X_test_flat.append(X_test_conv[k].reshape(1,lastNodes))
        for v in range(0, X_valid.shape[0]):
            X_valid_flat.append(X_valid_conv[v].flatten())
    else:
        for ib in range(0, X_train.shape[0]):
            X_train_flat.append(X_train_conv[ib].numpy().flatten())
            #X_train_flat.append(X_train_conv[ib].numpy().reshape(1,lastNodes))
        for k in range(0, X_test.shape[0]):
            X_test_flat.append(X_test_conv[k].numpy().flatten())
            #X_test_flat.append(X_test_conv[k].numpy().reshape(1,lastNodes))
        for v in range(0, X_valid.shape[0]):
            X_valid_flat.append(X_valid_conv[v].numpy().flatten())


    print(X_train_flat[0])
    print(len(X_train_flat[0]))
    print(len(X_train_flat))
    print(layertupples)
    convnode = len(X_train_flat[0])
    print(convnode,lastNodes)
    #input("check")
    print(type(X_train_flat[0]))
    print(X_train_flat[0].shape)
    print(convnode)
    #input("broken")

    X_train_flat = np.array(X_train_flat)
    X_test_flat = np.array(X_test_flat)
    X_valid_flat = np.array(X_valid_flat)

    return X_train_flat, X_test_flat, X_valid_flat, convnode



def sequentialTrain(X, Y, X_train, X_test, Y_train, Y_test, labels_name, test_input, outputType, result_name,
                    inferencefile, inf_filename, resultFileName):
    test_inputfile = np.array(test_input, np.int32)
    inf_filename = inf_filename

    allmodels = ['LR','SVR','RF','LIGHTGBM','XGBOOST']
    for an in range(1, 4):
        for bn in range(1, 11):
            annName = "ANN_%s_%s" % (an, bn)
            modelPath = "dataset_new_avg/elixir-modelsave/%s_%s_%s.h5" % (resultFileName, outputType, annName)
            if os.path.isfile(modelPath):
                allmodels.append(annName)
            else:
                print(annName)
                #input("wrong model append")
    for an in range(1, 4):
        for bn in range(1, 11):
            cnnName = "CNN_%s_%s" % (an, bn)
            modelPath = "dataset_new_avg/elixir-modelsave/%s_%s_%s.h5" % (resultFileName, outputType, cnnName)
            if os.path.isfile(modelPath):
                allmodels.append(cnnName)
                print(cnnName)
            else:
                print(cnnName)
                #input("wrong model append")


    outputset = Y
    singleOutputset = []
    for p in range(len(labels_name)):
        singleOutputset.append(outputset[labels_name[p]])

    inferenceall = []
    modelsall = []
    for tr in allmodels:
        if tr == 'LR':
            inference = linearRegression_sk(inferencefile, X_train, X_test, Y_train, Y_test, test_inputfile, outputType)
            # print(inference)
            predict = np.transpose(inference)
            # print(predict)
            inferenceall.append(predict)
        elif tr == 'SVR':
            print("svr")
            inference = lsvrModel(inferencefile, X_train, X_test, Y_train, Y_test, test_inputfile, outputType)
            # print(inference)
            predict = np.transpose(inference)
            # print(predict)
            inferenceall.append(predict)
        elif tr == 'RF':
            print("rf")
            inference = rforestModel(inferencefile, X_train, X_test, Y_train, Y_test, 10, test_inputfile, outputType)
            predict = np.transpose(inference)
            inferenceall.append(predict)
        elif tr == 'LIGHTGBM':
            print("lightgbm")
            predict = []
            for num in range(0, 4):
                e_train, e_test = train_test_split(singleOutputset[num], test_size=0.01, random_state=0)
                inference = lightgbm_model(X_train, X_test, e_train, e_test, test_input)
                predict.append(inference)
            # print(inference)
            inferenceall.append(predict)
        elif tr == 'XGBOOST':
            print("XGBOOST")
            predict = []
            for num in range(0, 4):
                e_train, e_test = train_test_split(singleOutputset[num], test_size=0.01, random_state=0)
                inference = XGBoost_model(X_train, X_test, e_train, e_test, test_input)
                # print(inference)
                predict.append(inference)
            inferenceall.append(predict)
        elif tr[0] == 'A':
            print("ANN")
            inference = annModel(X_train, X_test, Y_train, Y_test, test_inputfile, tr, outputType, resultFileName)
            # print(inference)
            predict = np.transpose(inference)
            # print(inference[num])
            # print("ann:",inference[num]

            inferenceall.append(predict)
        elif tr[0] == 'C':
            print("CNN")
            inference = cnnModel(X_train, X_test, Y_train, Y_test, test_inputfile, tr, outputType, resultFileName)
            # print(inference)
            predict = np.transpose(inference)
            # print(inference[num])
            # print("ann:",inference[num]

            inferenceall.append(predict)
        else:
            input("wrong model")

    outputName = [['avgSentMsg', 'avgSentByte', 'secMaxSendMsg', 'secMaxSendByte'],
                  ['avgRecvMsg', 'avgRecvByte', 'secMaxRecvMsg', 'secMaxRecvByte']]

    if outputType == "TX":
        for a in range(0, 4):
            wirteInference(test_inputfile, inferenceall, inferencefile, outputType, outputName[0][a], [],
                           resultFileName,
                           allmodels, result_name, labels_name, a)
    elif outputType == "RX":
        for a in range(0, 4):
            wirteInference(test_inputfile, inferenceall, inferencefile, outputType, outputName[1][a], [],
                           resultFileName,
                           allmodels, result_name, labels_name, a)


    # inferencefile.remove(inferencefile['Sheet'])
    inferencefile.save("dataset_new_avg/result/" + inf_filename)


def ann_randomnode(startNode, architecture, nodes, models, results, X_train, X_test, Y_train, Y_test, name_count,
                   rmse_set, outputType, inferencefile, test_inputfile, inf_filename, resultFileName):
    newNode = 8
    index_node = 0
    for i in range(1, 11):  # 레이어 갯수는 1~max레이어가 되도록
        architecture.initial_utnis([])
        annname = "ANN_" + str(name_count) + "_" + str(i)
        architecture.update_name(annname)
        newlayer = random.randint(1, 10)
        architecture.update_layers(newlayer)
        count = 1
        for j in range(0, newlayer):  # 레이어 갯수만큼 반복이 되야겠다. 레이어별 노드를 쌓는다.
            print("count: ", count)
            print("nodes(i): ", nodes)
            if newNode > 8:
                ranNode = 7
            else:
                if count == 1:
                    newNode = startNode[index_node]
                    if index_node == 7:
                        index_node = 0
                    else:
                        index_node = index_node + 1
                else:
                    ranNode = random.randint(4, 7)
                    print("ranNode: ", ranNode)
                    print("newNode: ", newNode)
                    while (newNode < (7 + nodes - ranNode)):
                        print("while: ", newNode)
                        if newNode == 4:
                            ranNode = 7
                            break;
                        else:
                            ranNode = random.randint(4, 7)
                            print("reNUMBER: ", ranNode)
                            print("nodes(i): ", nodes)

                    newNode = 7 + nodes - ranNode
            if newNode == 8:
                newNode = 7
            count = count + 1
            print("random_value for newNode : ", newNode)
            architecture.update_nodes(newNode)
        # initializer_index = random.randint(1,2)
        # activation_index = random.randint(1, 3)
        result, modelSummary, inference = annModel(architecture, X_train, X_test, Y_train, Y_test, test_inputfile,
                                                   annname, outputType, resultFileName)
        if result == None:
            continue
        wirteInference(test_inputfile, inference, inferencefile, outputType, annname)

        # inferencefile.save('/Users/yeonhooy/ccmodel/test_modelspace/' +"(PRE)" +inf_filename)
        architecture.update_modelSumary(modelSummary)
        print("ANN arc - ", architecture.get_arch())
        models.append(architecture.get_arch())
        results.append(result)
        print(result[0][4], result[1][4], result[2][4], result[3][4])
        if result[0][4] < rmse_set[0] and result[1][4] < rmse_set[1] and result[2][4] < rmse_set[2] and result[3][4] < \
                rmse_set[3]:
            print("best result!!")
            return 1
    return 2

def wirteInference(testdatafile, inference, resultfile, outputType, sheetname, supmodels, resultFileName, allmodels,
                   result_name, labels_name,outputNum):
    testdata = np.array(testdatafile, np.int32)
    print(testdata)

    print("inference--3D")
    print(inference)

    # ANSWER sheet for test_input
    #realdatafilename = "dataset_new_avg/inf_realvalue/new_score_csv/" + resultFileName + ".xlsx"
    realdatafilename = "dataset_new_avg/inf_realvalue/evaluation/" + resultFileName + ".xlsx"
    realdatafilename = "dataset_new_avg/inf_realvalue/dt_csv/" + resultFileName + ".xlsx"
    
    print(realdatafilename)
    real_wb = load_workbook(realdatafilename, data_only=True)
    #real_sheet = real_wb["Sheet"]
    #real_sheet = real_wb["Sheet1"]
    real_sheet = real_wb[resultFileName]
    # input("stop")

    #superemeloadfullpath = "dataset_new_avg/result/datcenter/" + result_name
    #supereme_wb = load_workbook(superemeloadfullpath, data_only=True)

    outminrmse = []
    indexrmse = []

    if outputType == 'TX':
        # 모델 당 sheet 생성
        sheet = resultfile.create_sheet(sheetname)
        #print(resultfile.get_sheet_names())
        # input("DD")

        if outputNum==0:
            sheet.cell(row=1, column=8).value = "avgSentMsg"
            # real value
            for r in range(0,len(inference[0][0])):
                sheet.cell(row=2+r,column=8).value = real_sheet.cell(row=2+r, column=9).value
            #inference value
            for m in range(0, len(allmodels)):
                for count in range(0,len(inference[0][0])):
                    print(m,count,allmodels[m])
                    sheet.cell(row=2+count,column=9+m).value=inference[m][0][count]
        elif outputNum==1:
            sheet.cell(row=1, column=8).value = "avgSentByte"
            for r in range(0,len(inference[0][0])):
                sheet.cell(row=2+r,column=8).value = real_sheet.cell(row=2+r, column=10).value
            # inference value
            for m in range(0, len(allmodels)):
                for count in range(0, len(inference[0][0])):
                    sheet.cell(row=2 + count, column=9 + m).value = inference[m][1][count]
        elif outputNum==2:
            sheet.cell(row=1, column=8).value = "secMaxSendMsg"
            for r in range(0,len(inference[0][0])):
                sheet.cell(row=2+r,column=8).value = real_sheet.cell(row=2+r, column=11).value
                # inference value
            for m in range(0, len(allmodels)):
                for count in range(0, len(inference[0][0])):
                    sheet.cell(row=2 + count, column=9 + m).value = inference[m][2][count]
        elif outputNum==3:
            sheet.cell(row=1, column=8).value = "secMaxSendByte"
            for r in range(0,len(inference[0][0])):
                sheet.cell(row=2+r,column=8).value = real_sheet.cell(row=2+r, column=12).value
            # inference value
            for m in range(0, len(allmodels)):
                for count in range(0, len(inference[0][0])):
                    sheet.cell(row=2 + count, column=9 + m).value = inference[m][3][count]

        sheet.cell(row=1, column=1).value = "Switch_e"
        sheet.cell(row=1, column=2).value = "Switch_c"
        sheet.cell(row=1, column=3).value = "host"
        sheet.cell(row=1, column=4).value = "connection"
        sheet.cell(row=1, column=5).value = "interval"
        sheet.cell(row=1, column=6).value = "link"
        sheet.cell(row=1, column=7).value = "hop"

        for t in range(0, len(testdata)):
            for p in range(0, 7):
                sheet.cell(row=2 + t, column=1 + p).value = testdata[t][p]
        for x in range(0, len(allmodels)):
            sheet.cell(row=1, column=9+x).value=str(allmodels[x])


    if outputType == 'RX':
        sheet = resultfile.create_sheet(sheetname)

        sheet.cell(row=1, column=1).value = "Switch_e"
        sheet.cell(row=1, column=2).value = "Switch_c"
        sheet.cell(row=1, column=3).value = "host"
        sheet.cell(row=1, column=4).value = "connection"
        sheet.cell(row=1, column=5).value = "interval"
        sheet.cell(row=1, column=6).value = "link"
        sheet.cell(row=1, column=7).value = "hop"

        for t in range(0, len(testdata)):
            for p in range(0, 7):
                sheet.cell(row=2 + t, column=1 + p).value = testdata[t][p]

        if outputNum==0:
            sheet.cell(row=1, column=8).value = "avgRecvMsg"
            for r in range(0,len(inference[0][0])):
                sheet.cell(row=2+r,column=8).value = real_sheet.cell(row=2+r, column=13).value
            # inference value
            for m in range(0, len(allmodels)):
                for count in range(0, len(inference[0][0])):
                    sheet.cell(row=2 + count, column=9 + m).value = inference[m][0][count]

        elif outputNum==1:
            sheet.cell(row=1, column=8).value = "avgRecvByte"
            for r in range(0,len(inference[0][0])):
                sheet.cell(row=2+r,column=8).value = real_sheet.cell(row=2+r, column=14).value
            # inference value
            for m in range(0, len(allmodels)):
                for count in range(0, len(inference[0][0])):
                    sheet.cell(row=2 + count, column=9 + m).value = inference[m][1][count]
        elif outputNum==2:
            sheet.cell(row=1, column=8).value = "secMaxRecvMsg"
            for r in range(0,len(inference[0][0])):
                sheet.cell(row=2+r,column=8).value = real_sheet.cell(row=2+r, column=15).value
            # inference value
            for m in range(0, len(allmodels)):
                for count in range(0, len(inference[0][0])):
                    sheet.cell(row=2 + count, column=9 + m).value = inference[m][2][count]
        elif outputNum==3:
            sheet.cell(row=1, column=8).value = "secMaxRecvByte"
            for r in range(0,len(inference[0][0])):
                sheet.cell(row=2+r,column=8).value = real_sheet.cell(row=2+r, column=16).value
            # inference value
            for m in range(0, len(allmodels)):
                for count in range(0, len(inference[0][0])):
                    sheet.cell(row=2 + count, column=9 + m).value = inference[m][3][count]
        for x in range(0, len(allmodels)):
            sheet.cell(row=1, column=9+x).value=str(allmodels[x])

def frequency_sort(data):
    rt_data = []
    for d, c in Counter(data).most_common():
        for i in range(c):
            rt_data.append(d)
    return rt_data


def wirteInferenceSingle(testdatafile, inference, resultfile, outputType, modelName):
    testdata = np.array(testdatafile, np.int32)
    print(testdata)

    if outputType == 'TX':
        # 모델 당 sheet 생성
        sheet = resultfile.create_sheet(modelName)
        sheet.cell(row=1, column=1).value = "Switch_e"
        sheet.cell(row=1, column=2).value = "Switch_c"
        sheet.cell(row=1, column=3).value = "host"
        sheet.cell(row=1, column=4).value = "connection"
        sheet.cell(row=1, column=5).value = "interval"
        sheet.cell(row=1, column=6).value = "link"
        sheet.cell(row=1, column=7).value = "hop"
        for t in range(0, len(testdata)):
            for p in range(0, 7):
                sheet.cell(row=2 + t, column=1 + p).value = testdata[t][p]

        sheet.cell(row=1, column=8).value = "avgSentMsg"
        sheet.cell(row=1, column=9).value = "avgSentByte"
        sheet.cell(row=1, column=10).value = "secMaxSendMsg"
        sheet.cell(row=1, column=11).value = "secMaxSendByte"
        for i in range(0, len(inference[0])):
            for k in range(0, 4):
                sheet.cell(row=2 + i, column=8 + k).value = inference[k][i]

    if outputType == 'RX':
        sheet = resultfile[modelName]
        sheet.cell(row=1, column=12).value = "avgRecvMsg"
        sheet.cell(row=1, column=13).value = "avgRecvByte"
        sheet.cell(row=1, column=14).value = "secMaxRecvMsg"
        sheet.cell(row=1, column=15).value = "secMaxRecvByte"
        for j in range(0, len(inference[0])):
            for l in range(0, 4):
                sheet.cell(row=2 + j, column=12 + l).value = inference[l][j]


def csvTrain_withoutKfold(result_excel, X, Y, result_name, worksheetName, test_input, outputType, inferencefile,
                          inf_filename, resultFileName):
    print("CSVTRAIN", X, Y)
    all_results = []
    X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.01, random_state=0)
    print(X_train, X_test, Y_train, Y_test)
    sequentialTrain(X, Y, X_train, X_test, Y_train, Y_test, worksheetName, test_input, outputType, result_name,
                    inferencefile, inf_filename, resultFileName)  # 3,3 --> (1,2,4) , (1,2,4)
    inferencefile.save("dataset_new_avg/result/datcenter/" + inf_filename)


def csvTrain(result_excel, X, Y, result_name, worksheetName):
    print("CSVTRAIN", X, Y)
    all_results = []
    # 1-3. Split Test / Train Data
    n = 5
    kf = KFold(n_splits=n)
    count = 1
    for train_index, test_index in kf.split(X, Y):
        # print("TRAIN:", train_index, "TEST:", test_index)
        X_train, X_test = X.iloc[train_index], X.iloc[test_index]
        Y_train, Y_test = Y.iloc[train_index], Y.iloc[test_index]
        # X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.2, random_state=0)
        print(X_train, X_test, Y_train, Y_test)

        # TODO: layer수, 노드 수가 변수이고, ann에서 다른 파라미터를 추가하기
        models, results = sequentialTrain(X, Y, X_train, X_test, Y_train, Y_test)  # 3,3 --> (1,2,4) , (1,2,4)
        print("COUNT: ", count)
        all_results.append(results)
        # print("All results", all_results)
        count = count + 1
    model_length = len(models)
    total_elements = model_length * 36
    tmp = np.arange(total_elements).reshape(model_length, 4, 9)
    # initial tmp[]
    for nx in range(0, model_length):
        for ny in range(0, 4):
            for nz in range(0, 9):
                tmp[nx][ny][nz] = 0
    tmp_n = np.arange(total_elements).reshape(model_length, 4, 9)
    # initial tmp[]
    for nx in range(0, model_length):
        for ny in range(0, 4):
            for nz in range(0, 9):
                tmp_n[nx][ny][nz] = 0
    print("init_TMP : ", tmp)
    print("LENGTH OF ALL_RESULTS : ", len(all_results))
    for i in range(len(all_results)):
        # tmp = tmp + all_results[i]
        print("Added all_results[i]", all_results[i])
        result_tmp = all_results[i]
        for a in range(model_length):
            for b in range(0, 4):
                for c in range(0, 9):
                    if result_tmp[a][b][c] != np.inf:
                        print(result_tmp[a][b][c], type(result_tmp[a][b][c]))
                        tmp[a][b][c] = tmp[a][b][c] + result_tmp[a][b][c]
                    elif result_tmp[a][b][c] == np.nan:
                        print("NAN", result_tmp[a][b][c])
                    else:
                        print("INF", result_tmp[a][b][c])

    print("tmp", tmp)
    tmp_n = np.round(tmp / n, 2)
    print("tmp_n", tmp_n)
    print("Modles: ", models)

    write_excel(result_excel, models, tmp_n, result_name, worksheetName)

    ##########################################
    ################ MAIN ####################
    ##########################################


def main():
    os.environ["CUDA_VISIBLE_DEVICES"] = "0,1,2,3"
    # file_name = input('파일 이름(확장자빼고)을 적고, 이 파일 이름대로 결과가 엑셀로 저장됩니다. >> ')
    # data_name = file_name + '.csv'
    # result_name = file_name + '.xlsx'
    pd.options.display.float_format = '{:.5f}'.format

    from tensorflow.python.client import device_lib


    # input("dd")

    # 0. get csv files in the folder.
    path_name = "dataset_new_avg"
    file_list = os.listdir(path_name)
    file_list_csv = [file for file in file_list if file.endswith(".csv")]
    file_list_csv.sort()
    print(file_list_csv)
    counter = 1

    # 1. get test data (for inference)
    test_path_name = "dataset_new_avg/inf_realvalue"
    testfile_list = os.listdir(test_path_name)
    testfile_list_csv = [file for file in testfile_list if file.endswith(".csv")]
    testfile_list_csv.sort()
    print(testfile_list_csv)

    #file_list_csv = ['p4_default.csv']
    #file_list_csv = ['onos_of10_default.csv', 'odl_of10_default', 'odl_of10_lldpstats', 'odl_of13_default', 'p4-total']
    for csvfile in file_list_csv:
        filenaming = csvfile.split('.')
        resultFileName = filenaming[0]
        print("Result file name : " + resultFileName)
        print("Now train: {}", path_name + "/" + csvfile)

        # 1. Read Data and remove outlier data
        print("++++++ Read Data ++++++")
        print(csvfile)
        dataset = pd.read_csv('dataset_new_avg/' + csvfile)
        print(dataset)
        print("origin dataset length", len(dataset))

        # 1-1 Outlier 탐색
        print("++++++ Outlier search ++++++")
        Outliers_to_drop = detect_outliers(dataset, 2, ["avgSentMsg"])
        dataset = dataset.drop(Outliers_to_drop, axis=0).reset_index(drop=True)

        Outliers_to_drop = detect_outliers(dataset, 2, ["avgSentByte"])
        dataset = dataset.drop(Outliers_to_drop, axis=0).reset_index(drop=True)

        Outliers_to_drop = detect_outliers(dataset, 2, ["secMaxSendMsg"])
        dataset = dataset.drop(Outliers_to_drop, axis=0).reset_index(drop=True)
        Outliers_to_drop = detect_outliers(dataset, 2, ["secMaxSendByte"])
        dataset = dataset.drop(Outliers_to_drop, axis=0).reset_index(drop=True)
        Outliers_to_drop = detect_outliers(dataset, 2, ["secMaxRecvMsg"])
        dataset = dataset.drop(Outliers_to_drop, axis=0).reset_index(drop=True)
        Outliers_to_drop = detect_outliers(dataset, 2, ["secMaxRecvByte"])
        dataset = dataset.drop(Outliers_to_drop, axis=0).reset_index(drop=True)
        print("after delete outlier dataset`s len", len(dataset))
        dataseted = np.array(dataset)
        lendataset = len(dataset) - 2
        print(dataseted[0])
        # input("DD")

        X = dataset.drop(
            labels=['time', 'avgSentMsg', 'avgSentByte', 'secMaxSendMsg', 'secMaxSendByte', 'avgRecvMsg', 'avgRecvByte',
                    'secMaxRecvMsg', 'secMaxRecvByte'], axis=1)  # 7개의 Input data / label
        # output for TX / RX
        y_1 = dataset[['avgSentMsg', 'avgSentByte', 'secMaxSendMsg', 'secMaxSendByte']]
        y_2 = dataset[['avgRecvMsg', 'avgRecvByte', 'secMaxRecvMsg', 'secMaxRecvByte']]

        # output for msg / byte
        # y_3 = dataset[['avgSentMsg', 'secMaxSendMsg', 'avgRecvMsg', 'secMaxRecvMsg']]
        # y_4 = dataset[['avgSentByte', 'secMaxSendByte', 'avgRecvByte', 'secMaxRecvByte']]

        # y = [y_1, y_2, y_3, y_4]
        y = [y_1, y_2]

        print(X)
        print(y_1)
        print(y_2)
        # print(y_3)
        # print(y_4)

        # pathfullName = 'dataset_new_avg/inf_realvalue/'+resultFileName + '.xlsx'
        # print("test file name : " + pathfullName)
        #
        # testdatasetp =
        # _workbook(pathfullName, data_only=True)
        # testdataset = testdatasetp[resultFileName]
        # print(testdataset)

        # trainfullpath = 'dataset_new_avg/train_xlsx/'+resultFileName + '.xlsx'
        # traindataset_part = load_workbook(pathfullName, data_only=True)
        # traindataset_part = traindataset_part[resultFileName]

        # todo : This code is for add trainingset(72) --> scoring set / for test
        # for add in range(0,72):
        #     next_row=testdataset.max_row+1
        #     nowrow=random.randint(1,lendataset)
        #     for now in range(0,16):
        #         testdataset.cell(row=next_row,column=now+1).value = dataseted[nowrow][now]
        # savefullName='dataset_new_avg/inf_realvalue/csv/'+resultFileName + '.xlsx'
        # testdatasetp.save(savefullName)
        # input("save csv file")

        # QUESTION Sheet
        #testdatasetw = pd.read_csv('dataset_new_avg/inf_realvalue/new_score_csv/' + resultFileName + '.csv')
        #testdatasetw = pd.read_csv('dataset_new_avg/inf_realvalue/evaluation/' + resultFileName + '.csv')
        testdatasetw = pd.read_csv('dataset_new_avg/inf_realvalue/dt_csv/' + resultFileName + '.csv')

        # --> test input is for test
        test_input = testdatasetw[['switch_e', 'switch_c', 'host', 'connection', 'interval', 'link', 'hop']]
        # --> test input is for test

        outputTypeNum = ["TX", "RX", "MSG", "Byte"]
        resultfilename = ["", "", "", ""]
        outputTitles = [['avgSentMsg', 'avgSentByte', 'secMaxSendMsg', 'secMaxSendByte'],
                        ['avgRecvMsg', 'avgRecvByte', 'secMaxRecvMsg', 'secMaxRecvByte'],
                        ['avgSentMsg', 'secMaxSendMsg', 'avgRecvMsg', 'secMaxRecvMsg'],
                        ['avgSentByte', 'secMaxSendByte', 'avgRecvByte', 'secMaxRecvByte']]
        print("OutputitleTest", outputTitles[0])

        # inference file 생성
        import openpyxl
        inferencefile = openpyxl.Workbook()
        test_inputfile = np.array(test_input, np.int32)
        inf_filename = resultFileName + "_inference.xlsx"
        inferencefile.save("dataset_new_avg/result/" + inf_filename)

        for num in range(0, 2):
            result_excel = Workbook()
            resultfilename[num] = resultFileName + "_result_" + outputTypeNum[num] + ".xlsx"
            print("resultNAME: ", resultfilename[num])
            # csvTrain(result_excel, X, y[num], resultfilename[num],outputTitles[num])
            csvTrain_withoutKfold(result_excel, X, y[num], resultfilename[num], outputTitles[num], test_input,
                                  outputTypeNum[num], inferencefile, inf_filename, resultFileName)
            # result_excel.remove(result_excel['Sheet'])
            # result_excel.save("dataset_new_avg/result/datcenter/" + resultfilename[num])
            # inferencefile.remove(inferencefile['Sheet'])
            inferencefile.save("dataset_new_avg/result/datcenter/" + inf_filename)
            counter = counter + 1
            print("COUNTER: ", counter)
    print("Finished!", outputTypeNum[num])


main()








