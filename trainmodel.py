# Copyright 2023 by Yeonho Yoo, Operating systems lab (Korea university, Seoul).
# All rights reserved.
# This file is part of Machine Learning-based Prediction Models for Control Traffic in SDN Systems (Elixr),
# and is released under the "MIT License Agreement". Please see the LICENSE
# file that should have been included as part of this package.

from __future__ import absolute_import, division, print_function, unicode_literals, unicode_literals

import math
import pathlib
import string
from ast import literal_eval
import matplotlib.pyplot as plt


import numpy as np
import pandas as pd
import tensorflow as tf
import tensorflow.compat.v1 as v1
from tensorflow import keras
from tensorflow.keras.regularizers import l1,l2
from tensorflow.keras.optimizers import SGD,Adam,RMSprop
from tensorflow.keras.layers import Dropout
from tensorflow.keras.models import Sequential
from tensorflow.keras import layers
from tensorflow.keras.layers import Dense
from tensorflow.keras.layers import Flatten
from tensorflow.keras.layers import Conv1D
from tensorflow.keras.layers import MaxPooling1D


from sklearn.pipeline import make_pipeline,Pipeline
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
from sklearn.svm import LinearSVR
from sklearn.multioutput import MultiOutputRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import mean_absolute_error,accuracy_score
from sklearn.preprocessing import StandardScaler,MinMaxScaler, RobustScaler
from sklearn.model_selection import KFold
from sklearn.metrics import r2_score

import lightgbm as lgb

import xgboost as xgb

import random
import sys
from itertools import combinations

from scipy import stats


# Ignore all GPUs, tf random forest does not benefit from it.
import os
import shutil
import datetime
#os.environ["TF_CPP_MIN_LOG_LEVEL"] = "2"
# os.environ["CUDA_VISIBLE_DEVICES"] = ""
# test for 10-05

from openpyxl import Workbook


class PrintDot(keras.callbacks.Callback):
  def on_epoch_end(self, epoch, logs):
    if epoch % 100 == 0: print('')
    print('~', end='')  #print ~ when 1 epoch end

# 1.functions
# <1. detect outlier and remove outliers>
def detect_outliers(df, n, features):
    outlier_indices=[]
    print("df",df)
    for col in features:
        print("col",col) #[Output]
        Q1 = np.percentile(df[col], 25)
        Q3 = np.percentile(df[col], 75)
        print("Q1",Q1)
        print("Q3",Q3)
        IQR = Q3 - Q1
        outlier_step = 1.5 * IQR
        minQ = Q1 - outlier_step
        maxQ = Q3 + outlier_step
        print("minQ",minQ)
        print("maxQ",maxQ)
        outlier_list_col = df[(df[col] < minQ) | (df[col] > maxQ)].index
        outlier_indices.extend(outlier_list_col)
        print("Outlier`s len",len(outlier_indices))
        print("Outlier",outlier_indices)

    return outlier_indices
def detect_outliers_delete(df, n, features):
    outlier_indices=[]
    print("df",df)
    for col in features:
        print("col",col) #[Output]
        Q1 = np.percentile(df[col], 25)
        Q3 = np.percentile(df[col], 75)
        print("Q1",Q1)
        print("Q3",Q3)
        IQR = Q3 - Q1
        outlier_step = 1.5 * IQR
        minQ = Q1 - outlier_step
        maxQ = Q3 + outlier_step
        print("minQ",minQ)
        print("maxQ",maxQ)
        outlier_list_col = df[(df[col] < minQ) | (df[col] > maxQ)].index
        outlier_indices.extend(outlier_list_col)
        print("Outlier`s len",len(outlier_indices))
        print("Outlier",outlier_indices)
        for i in range(len(outlier_indices)):
            df = df.drop(df.index[outlier_indices[i]])
        print ("new df ",df)

    return df

def print_accuracy_all_single(test_predictions,y_test):
    #Array of resutls
    result = [0,0,0,0,0,0,0,0,0]
    # Accuracy Measurement
    dot_70 = 0
    dot_80 = 0
    dot_90 = 0
    abserror = np.abs((test_predictions - y_test.array))
    errorrate = abserror / (y_test.array) * 100
    maperate = np.abs((test_predictions-y_test.array)/y_test.array) * 100
    accrate = 100 - errorrate

    for u in accrate:
        if (u >= 70 and u <= 100):
            dot_70 = dot_70 + 1
    result[0] = round(dot_70 / len(accrate) * 100 ,3)
    print("OVER 70: ")
    print(result[0])

    for q in accrate:
        if (q >= 80 and q <= 100):
            dot_80 = dot_80 + 1
    result[1] = round(dot_80 / len(accrate) * 100 ,3)
    print("OVER 80: ")
    print(result[1])

    for z in accrate:
        if (z >= 90 and z <= 100):
            dot_90 = dot_90 + 1
    result[2] = round(dot_90 / len(accrate) * 100 ,3)
    print("OVER 90: ")
    print(result[2])

    # MSE
    n_mse = len(accrate)
    square_mse = np.square(abserror)
    result[3] = round(np.sum(square_mse) / n_mse ,3)
    print("MSE: ")
    print(result[3])

    #RMSE
    n_rmse = len(accrate)
    mse = np.square(abserror)
    smse = np.sum(mse) / n_rmse
    result[4] = round(np.sqrt(smse), 3)
    print("RMSE: ")
    print(result[4])

    # MAE
    n_mae = len(accrate)
    result[5] = round(np.sum(abserror) / n_mae, 3)
    print("Mae: ")
    print(result[5])

    # MAPE
    n_mape = len(accrate)
    result[6] = round(np.sum(maperate) / n_mape, 3)
    print("MAPE: ")
    print(result[6])

    #Pearson R
    # val_sum = np.sum(test_predictions)
    # array_has_stranges = np.isnan(val_sum) + np.isinf(val_sum)
    # val_sum2 = np.sum(y_test)
    # arrayhas_stranges2 = np.isnan(val_sum2) + np.isinf(val_sum2)

    if any(np.isnan(val) for val in test_predictions.astype(np.float32)) or any(np.isnan(val) for val in y_test.astype(np.float32)) or any(np.isinf(val) for val in test_predictions.astype(np.float32)) or any(np.isinf(val) for val in y_test.astype(np.float32)):
        # test_predictions.fillna(test_predictions.mean())
        # y_test.fillna(y_test.mean())
        result[7] = 0
    else:
        pearsonr = stats.pearsonr(y_test, test_predictions)
        result[7] = round(pearsonr[0],3)

    print("Pearson R: ")
    print(result[7])



    #R2

    if any(np.isnan(val) for val in test_predictions.astype(np.float32)) or any(np.isnan(val) for val in y_test.astype(np.float32)) or any(np.isinf(val) for val in test_predictions.astype(np.float32)) or any(np.isinf(val) for val in y_test.astype(np.float32)):
        # test_predictions.fillna(test_predictions.mean())
        # y_test.fillna(y_test.mean())
        result[8] = 0
    else:
        r2cal = r2_score(y_test, test_predictions)
        result[8] = round(r2cal,3)

    print("R2: ")
    print(result[8])



    return result

# <2. print result with 5 metrics>
def print_accuracy_all(test_predictions,y_test,modelName):
    result_1 = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    result_2 = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    result_3 = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    result_4 = [0, 0, 0, 0, 0, 0, 0, 0, 0]

    result = {
        'Recv_bw' : result_1,
        'Sent_bw' : result_2,
        'Max_sent' : result_3,
        'Max_recv' : result_4
    }

    # Accuracy Measurement
    dot_70 = 0
    dot_80 = 0
    dot_90 = 0

    abserror = np.abs((test_predictions - y_test))
    errorrate = abserror / (y_test) * 100
    maperate = np.abs((test_predictions-y_test)/y_test) * 100
    accrate = 100 - errorrate
    acc_array = np.array(accrate)

    print(test_predictions)
    print(y_test)

    predict_col_1 = np.array(test_predictions).T[0]
    predict_col_2 = np.array(test_predictions).T[1]
    predict_col_3 = np.array(test_predictions).T[2]
    predict_col_4 = np.array(test_predictions).T[3]

    #print(predict_col_1,predict_col_2,predict_col_3,predict_col_4)

    test_col_1 = y_test.iloc[:,0]
    test_col_2 = y_test.iloc[:,1]
    test_col_3 = y_test.iloc[:,2]
    test_col_4 = y_test.iloc[:,3]


    test_col_1 = np.array(test_col_1).flatten()
    test_col_2 = np.array(test_col_2).flatten()
    test_col_3 = np.array(test_col_3).flatten()
    test_col_4 = np.array(test_col_4).flatten()

    acc_col_1 = accrate.iloc[:,0]
    acc_col_2 = accrate.iloc[:,1]
    acc_col_3 = accrate.iloc[:,2]
    acc_col_4 = accrate.iloc[:,3]

    acc_col_1 = np.array(acc_col_1).flatten()
    acc_col_2 = np.array(acc_col_2).flatten()
    acc_col_3 = np.array(acc_col_3).flatten()
    acc_col_4 = np.array(acc_col_4).flatten()

    print(modelName)
    result = [calculate_accrate('Sent_bw',predict_col_1,test_col_1,acc_col_1,result_1),calculate_accrate('Recv_bw',predict_col_2,test_col_2,acc_col_2,result_2),
              calculate_accrate('Max_sent',predict_col_3,test_col_3,acc_col_3,result_3),calculate_accrate('Max_recv',predict_col_4,test_col_4,acc_col_4,result_4)]


    print(result)
    return result

def calculate_accrate(name,predict,test,accrate,result):
    #print(name)
    dot_70=0
    dot_80 = 0
    dot_90 = 0
    for u in accrate:
        if (u >= 70 and u <= 100):
            dot_70 = dot_70 + 1
    result[0] = round(dot_70 / len(accrate) * 100,3)
    # print("OVER 70: ")
    # print(result[0])

    for q in accrate:
        if (q >= 80 and q <= 100):
            dot_80 = dot_80 + 1
    result[1] = round(dot_80 / len(accrate) * 100,3)
    # print("OVER 80: ")
    # print(result[1])

    for z in accrate:
        if (z >= 90 and z <= 100):
            dot_90 = dot_90 + 1
    result[2] = round(dot_90 / len(accrate) * 100,3)
    # print("OVER 90: ")
    # print(result[2])


    abserror = np.abs((predict - test))
    maperate = np.abs((predict - test) / test) * 100

    # MSE
    n_mse = len(test)
    square_mse = np.square(abserror)
    result[3] = round(np.sum(square_mse) / n_mse,3)
    # print("MSE: ")
    # print(result[3])

    #RMSE
    n_rmse = len(test)
    mse = np.square(abserror)
    smse = np.sum(mse) / n_rmse
    result[4] = round(np.sqrt(smse),3)
    #print("RMSE: ")
    #print(result[4])

    # MAE
    n_mae = len(test)
    result[5] = round(np.sum(abserror) / n_mae,3)
    #print("Mae: ")
    #print(result[5])

    # MAPE
    n_mape = len(test)
    result[6] = round(np.sum(maperate) / n_mape,3)
    #print("MAPE: ")
    #print(result[6])

    # Pearson R
    if any(np.isnan(val) for val in predict.astype(np.float32)) or any(
            np.isnan(val) for val in test.astype(np.float32)) or any(
            np.isinf(val) for val in predict.astype(np.float32)) or any(
            np.isinf(val) for val in test.astype(np.float32)):
        # test_predictions.fillna(test_predictions.mean())
        # y_test.fillna(y_test.mean())
        #result[7] = 'NA'
        result[7] = 0
    else:
        pearsonr = stats.pearsonr(test, predict)
        result[7] =round(pearsonr[0],3)

    # print("Pearson R: ")
    # print(result[7])

    # R2
    if any(np.isnan(val) for val in predict.astype(np.float32)) or any(
            np.isnan(val) for val in test.astype(np.float32)) or any(
            np.isinf(val) for val in predict.astype(np.float32)) or any(
            np.isinf(val) for val in test.astype(np.float32)):
        # test_predictions.fillna(test_predictions.mean())
        # y_test.fillna(y_test.mean())
        #result[8] = 'NA'
        result[8] = 0

    else:
        r2cal = r2_score(test, predict)
        result[8] = round(r2cal,3)
        # print("R2: ")
        # print(result[8])

    return result


# <3. show predict accuracy graph>
def plot_history(history):
  hist = pd.DataFrame(history.history)
  hist['epoch'] = history.epoch

  plt.figure(figsize=(6,9))

  plt.subplot(2,1,1)
  plt.xlabel('Epoch')
  plt.ylabel('Mean Abs Error [Output]')
  plt.plot(hist['epoch'], hist['mean_absolute_error'],
           label='Train Error')
  plt.plot(hist['epoch'], hist['val_mean_absolute_error'],
           label = 'Val Error')
  plt.ylim([0,5])
  plt.legend()

  plt.subplot(2,1,2)
  plt.xlabel('Epoch')
  plt.ylabel('Mean Square Error [Output^2$]')
  plt.plot(hist['epoch'], hist['mean_squared_error'],
           label='Train Error')
  plt.plot(hist['epoch'], hist['val_mean_squared_error'],
           label = 'Val Error')
  plt.ylim([0,20])
  plt.legend()
  plt.show()


# 4. Write result to excel file

def write_excel(result_excel, models, results,result_name,outputTitles,rmsemodel,minrmse):
    # outputTitles = ['Sent_bw','Recv_bw', 'Max_sent', 'Max_recv']
    print("NOW SAVE ",result_name,results)
    print("TEST: ",results[0][0][0])
    for o in range(0,4):
        eIndex = o+1
        wTitle = outputTitles[o]
        result_excel.create_sheet(index = eIndex, title=wTitle)
        ws_1 = result_excel[wTitle]
        ws_1.cell(row=1, column=1).value = 'Model'
        ws_1.cell(row=1, column=2).value = 'Layer'
        ws_1.cell(row=1, column=3).value = 'Nodes'
        ws_1.cell(row=1, column=4).value = 'Over 70'
        ws_1.cell(row=1, column=5).value = 'Over 80'
        ws_1.cell(row=1, column=6).value = 'Over 90'
        ws_1.cell(row=1, column=7).value = 'MSE'
        ws_1.cell(row=1, column=8).value = 'RMSE'
        ws_1.cell(row=1, column=9).value = 'MAE'
        ws_1.cell(row=1, column=10).value = 'MAPE'
        ws_1.cell(row=1, column=11).value = 'Pearson R'
        ws_1.cell(row=1, column=12).value = 'Pearson R2'
        ws_1.cell(row=1, column=13).value = 'Design'
        counter = 2

        model_len = len(models)
        print("LENMODEL: ",model_len)
        for i in range(len(models)):
            print("MODEL_SUMMARY: ",str(models[i][4]))
            ws_1.cell(row=counter, column=1).value = models[i][0]
            ws_1.cell(row=counter, column=2).value = models[i][1]
            #ws_1.cell(row=counter, column=3).value = models[i][2]
            ws_1.cell(row=counter, column=3).value = str(models[i][3])
            for j in range(0,9):
                colj=4+j
                print(i,o,j)
                ws_1.cell(row=counter, column=colj).value = results[i][o][j]
            ws_1.cell(row=counter, column=13).value = str(models[i][4])
            counter = counter + 1
        ws_1.cell(row=40,column=1).value = rmsemodel[o]
        ws_1.cell(row=40, column=2).value = minrmse[o]


#2. Models
# 2-1) <Linear Regresssion>  - /multiple output regression /
def lrModel(X_train,X_test,y_train,y_test):
    # 1.minmax Sclaer
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)

    print(X_train,len(X_train))
    print(X_test,len(X_test))


    #2. Build linearRegresion
    model = Sequential()
    model.add(Dense(4, activation='linear', input_dim=X_train.shape[1]))
    #model.add(Dense(4))

    optimizer = tf.train.GradientDescentOptimizer(learning_rate=0.001)
    # optimizer = tf.keras.optimizers.RMSprop(0.001)

    model.compile(loss='mean_squared_error',
                    optimizer=optimizer,
                    metrics=['mean_absolute_error', 'mean_squared_error'])

    print("<<<<<LR>>>>>")
    model.summary()
    early_stop = keras.callbacks.EarlyStopping(monitor='val_loss', patience=10)

    model.fit(
        X_train, y_train.to_numpy(),
        epochs=1000, batch_size=10, validation_split=0.2, verbose=0,
        callbacks=[early_stop, PrintDot()])

    #loss, mean_absolute_error, mean_squared_error = lRmodel.evaluate(X_test, y_test.to_numpy(), verbose=2)

    test_predictions = model.predict(X_test)
    print(X_test)
    print("***********************************************************************************************************************************",test_predictions)
    result = print_accuracy_all(test_predictions, y_test,'Linear Regression')

    # Save Trained Model
    #model.save('alpha_LR.h5')

    return result

def linearRegression_sk(X_train,X_test,y_train,y_test,outputType, valid_input, valid_output):
    # Scaling
    # 1.minmax Sclaer
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    #ptest_input = scaler.transform(test_input)
    ptest_input = scaler.transform(valid_input)


    model = make_pipeline(RobustScaler(), MultiOutputRegressor(LinearRegression()))
    model.fit(X_train, y_train)

    test_predictions = model.predict(X_test)
    result = print_accuracy_all(test_predictions, y_test,"Linear regression with sk")
    valid = model.predict(ptest_input)
    #wirteInference(test_input,inference,inferencefile,outputType,"LR")

    return result,valid

# 2) <Linear SVR>  / single output - todo : sinlge --> mutliple outptut / (done)
def lsvrModel(X_train,X_test,y_train,y_test,outputType, valid_input, valid_output):
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    ptest_input = scaler.transform(valid_input)
    #test_input = scaler.transform(test_input)
    model =  make_pipeline(RobustScaler(), MultiOutputRegressor(LinearSVR(random_state=0, tol=1e-05, max_iter=5000)))
    #model.fit(X_train, y_train.values.ravel())
    model.fit(X_train,y_train)
    test_predictions = model.predict(X_test)
    result = print_accuracy_all(test_predictions, y_test,"Linear SVR")
    valid = model.predict(ptest_input)
    #wirteInference(test_input, inference, inferencefile, outputType, "SVR")

    # Save Trained Model
    #model.save('alpha_SVR.h5')

    return result, valid

# 3) <ForestRegression>
def rforestModel(X_train, X_test, y_train, y_test, outputType, valid_input, valid_output,tree):
    # Scaling
    # 1.minmax Scaler
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    ptest_input = scaler.transform(valid_input)

    #2. build model
    model = RandomForestRegressor(n_estimators=tree)
    model.fit(X_train, y_train)

    #3.Predict & Test the Model ############
    real_output = np.array(X_test)
    inputs = real_output
    predicted_output_3 = model.predict(inputs)
    #test_predictions_3 = predicted_output_3.flatten()
    result = print_accuracy_all(predicted_output_3, y_test,"Random Forest")

    #inference
    test_output = np.array(ptest_input)
    inference = model.predict(test_output)

    #wirteInference(test_input, inference, inferencefile, outputType, "RF")

    # Save Trained Model
    #model.save('alpha_RF.h5')

    return result, inference


def lightgbm_model(X_train, X_test, y_train, y_test, valid_input, valid_output):
    print("*******lightgbm**********")
    print(X_train.shape, X_test.shape, y_train.shape, y_test.shape)
    X_train, X_valid, y_train, y_valid = train_test_split(X_train, y_train, test_size=0.1, random_state=20)
    train_data = lgb.Dataset(X_train, label=y_train)
    valid_data = lgb.Dataset(X_valid, label=y_valid)
    #val_dta = lgb.Dataset(valid_input)
    print(y_valid)

    params = {'learning_rate': 0.01,
              'max_depth': 16,
              'boosting': 'gbdt',
              'objective': 'regression',
              'metric': ['mse','mae','mape'],
              'is_training_metric': True,
              'num_leaves': 144,
              'feature_fraction': 0.9,
              'bagging_fraction': 0.7,
              'bagging_freq': 5,
              'seed':2018}

    #num_round = 10

    model = lgb.train(params, train_data, 1000, valid_sets=[valid_data])
    # delete early_stopping
    # model = lgb.train(params, train_data, 1000, valid_sets=[valid_data], early_stopping_rounds=100)

    predict_test = model.predict(X_test)

    inference = model.predict(valid_input)

    return print_accuracy_all_single(predict_test.flatten(), y_test), inference


def XGBoost_model(X_train, X_test, y_train, y_test, valid_input, valid_output):
    print("*******XGBOOST MODEL*******")
    train_data = xgb.DMatrix(data=X_train, label=y_train)
    test_data = xgb.DMatrix(data=X_test, label=y_test)
    inf_dta = xgb.DMatrix(data=valid_input)
    print("DDDDDDDDDDDDDDDDDDD")
    print(test_data,inf_dta)
    print(X_test,valid_input)

    param = {'max_depth': 8, 'objective': 'reg:squarederror', 'eval_metric': ['rmse', 'mae']}
    num_round = 10

    #BASIC MODEL
    bst = xgb.train(param, train_data, num_round)
    # make prediction
    preds = bst.predict(test_data)

    inference = bst.predict(inf_dta)

    #RF MODEL
#    bst = xgb.XGBRFRegressor(random_state=42).fit(X_train, y_train)
#    preds = bst.predict(X_test)

    return print_accuracy_all_single(preds.flatten(), y_test), inference


# 4) <ANN>
class Architecture:
    def __init__(self, name, layer, nodes, design):
        self.name = name
        self.layer = layer
        self.nodes = nodes
        self.units = []
        self.modelSummary = []
        self.design = design
    def initial_utnis(self, units):
        self.units = units
    def update_name(self, name):
        self.name = name
        #self.units = self.design.nodesUnit(self.layer, self.nodes)
    def update_nodes(self, nodes):
        self.nodes = nodes
        self.units = self.design.nodesUnit(self.layer, self.nodes, self.units)
    def update_units(self, design_type):
        self.design.type = design_type
        self.units = self.design.nodesUnit(self.layer, self.nodes, self.units)
    def update_layers(self, layer):
        self.layer = layer
        self.units = self.design.nodesUnit(layer, 0, self.units)
    def update_modelSumary(self, modelSummary):
        self.modelSummary = modelSummary
    def get_arch(self):
        tmp = [self.name, self.layer, self.nodes, self.units, self.modelSummary]
        #tmp = tmp + self.design.get_design()
        return tmp

class Design:
    def __init__(self):
        self.type
        self.dropout
        self.regularizer
        self.initializer
        self.lr
        self.beta_1
        self.beta_2
        self.epsilon
        self.decay
        self.optimizer
    def nodesUnit(self, layer, nodes, units):
        if self.type == 'Default':
            print("info-",layer,nodes, units)
            if nodes != 0:
                units.append(nodes)
            print("(Defalut)First line Tmp : ", units)
            #tmp.append(4)  #for output node?
            print("(Defalut) Tmp : ", units )
            return units
        if self.type == 'First_half':
            tmp = [i for i in range(4,nodes+1)]
            tmp = tmp + [nodes for i in range(layer-nodes-1)]
            print("(FirstHalf)First line Tmp : ", tmp)
            #tmp.append(4)  # for output node?
            print("(FirstHalf)Second line Tmp : ", tmp)
            return tmp


    def get_design(self):
        return [self.type, self.dropout, self.regularizer,
                self.initializer, self.lr, self.beta_1,
                self.beta_2, self.epsilon, self.decay, self.optimizer]

class Default(Design):
    def __init__(self):
        self.name = 'ANN'
        self.type = 'Default'
        self.dropout = 0.2
        self.regularizer = l2(0.001)
        self.initializer = ['he_normal','he_uniform']
        self.activations = ['elu','relu','linear']
        self.lossfunctions = ['msle','mse','mae','mape']
        self.lr = 0.001
        self.beta_1 = 0.9
        self.beta_2 = 0.999
        self.epsilon = 1e-08
        self.decay = 0.00000001
        # self.optimizer = Adam(lr=self.lr, beta_1=self.beta_1, beta_2=self.beta_2, epsilon=self.epsilon, decay=self.decay)
        self.optimizer = Adam(lr=self.lr, beta_1=self.beta_1, beta_2=self.beta_2, epsilon=self.epsilon)

def annModel(architecture, X_train, X_test, y_train, y_test, annname,outputType,resultFileName, valid_input,valid_output):
    # checkpoint
    # 0.save model
    checkpoint_path = "dataset_new_avg/elixir-modelsave/annmodel_cp/%s.ckpt" % annname
    checkpoint_dir = os.path.dirname(checkpoint_path)
    cp_callback = tf.keras.callbacks.ModelCheckpoint(filepath=checkpoint_path, save_weights_only=True, verbose=1)

    # Scaling
    # 1.minmax sclaer
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    modelSummary = []

    # function api
    # 1) define input node / output node
    design = architecture.design
    inputs = keras.Input(shape=X_train.shape[1])
    ran_initializer_output = random.choice(design.initializer)
    outputlayer = layers.Dense(4)

    # function for dense
    def funden(dense, x):
        return dense(x)

    # 2) create new node in layer for ANN network

    if architecture.layer == 1:
        ran_initializer = random.choice(design.initializer)
        ran_activation = random.choice(design.activations)
        # first hidden node
        dense = layers.Dense(architecture.units[0], kernel_initializer=ran_initializer, activation=ran_activation,
                             kernel_regularizer=design.regularizer)
        modelSummary.append(architecture.units[0])
        modelSummary.append(ran_initializer)
        modelSummary.append(ran_activation)
        finaloutput = funden(outputlayer, funden(dense, inputs))
    if architecture.layer > 1:
        denselist = []
        outputlist = []
        finaloutputs = [inputs]

        # Adding the input layer
        ran_initializer = random.choice(design.initializer)
        ran_activation = random.choice(design.activations)

        dense = layers.Dense(architecture.units[0], kernel_initializer=ran_initializer, activation=ran_activation,
                             kernel_regularizer=design.regularizer)
        outputlist.append(funden(dense, inputs))

        modelSummary.append(architecture.units[0])
        modelSummary.append(ran_initializer)
        modelSummary.append(ran_activation)

        for i in range(architecture.layer - 1):
            ran_initializer = random.choice(design.initializer)
            ran_activation = random.choice(design.activations)
            denselist.append(
                layers.Dense(architecture.units[i + 1], kernel_initializer=ran_initializer, activation=ran_activation,
                             kernel_regularizer=design.regularizer))
            modelSummary.append(architecture.units[i + 1])
            modelSummary.append(ran_initializer)
            modelSummary.append(ran_activation)
        print(len(denselist))
        for x in range(len(denselist)):
            finaloutputs.append(funden(denselist[x], finaloutputs[x]))
            output = finaloutputs[x + 1]
            print(output, x)

        # Adding the output layer
        finaloutput = funden(outputlayer, output)
        modelSummary.append("Output(4)")
        modelSummary.append(ran_initializer_output)

    model = keras.Model(inputs=inputs, outputs=finaloutput, name=annname)

    # Compiling the ANN
    ran_lossfunc = random.choice(design.lossfunctions)
    modelSummary.append(ran_lossfunc)
    model.compile(optimizer=design.optimizer,
                  loss=ran_lossfunc, metrics=['mse', 'accuracy'])

    print("<<<<<<ANN_space : %s>>>>>>>>>>>" % annname)
    model.summary()

    early_stop = tf.keras.callbacks.EarlyStopping(monitor='val_loss', mode='min', patience=10)
    # Fitting the ANN to the Training set
    # regressor.fit(X_train, y_train, batch_size = 10, verbose=1, epochs = 1000, validation_split=0.2)
    start_time = datetime.datetime.now()
    print(X_train[0])

    model.fit(X_train, y_train, batch_size=32, verbose=1, epochs=1000, validation_split=0.15, callbacks=[early_stop])
    end_time = datetime.datetime.now()
    train_time = (end_time - start_time).total_seconds()
    model_name = "%s_%s_%s" % (resultFileName, outputType, annname)
    string_time = model_name + " : " + str(train_time) + "\n"

    print(start_time, end_time, train_time, string_time)
    # input("check")


    file = open("dataset_new_avg/elixir-modelsave/train_time.txt", "a")
    file.write(string_time)
    file.close()


    # save model
    savemodel_name = "dataset_new_avg/elixir-modelsave/%s_%s_%s.h5" % (resultFileName, outputType, annname)
    model.save(savemodel_name)

    ############ Predict & Test the Model ############
    real_output = np.array(X_test)
    inputs = real_output
    predicted_output = model.predict(inputs)
    outputs = np.array(y_test)
    print("ANN predict output")
    print(predicted_output)

    valid_inference = model.predict(valid_input)
    print(valid_inference)

    valid_output = np.array(valid_output)
    rmse_1 = np.square(valid_inference-valid_output)
    print(rmse_1)


    validsetlen = len(valid_output)
    rmse_annset = []
    for x in range(0,4):
        rmse_1_sum=0
        for y in range(0,validsetlen):
            rmse_1_sum=rmse_1_sum+rmse_1[y][x]
        rmse_annset.append(math.sqrt(rmse_1_sum)/validsetlen)
    print(rmse_annset)

    del model
    # print accuracy
    result = print_accuracy_all(predicted_output, y_test, "ANN")
    return result, modelSummary, rmse_annset

def cnnModel(architecture, X_train, X_test, y_train, y_test, cnnname,outputType,resultFileName, valid_input,valid_output):
    # convolution layer + ann
    # checkpoint
    # 0.save model
    checkpoint_path = "dataset_new_avg/elixir-modelsave/cnnmodel_cp/%s.ckpt" % cnnname
    checkpoint_dir = os.path.dirname(checkpoint_path)
    cp_callback = tf.keras.callbacks.ModelCheckpoint(filepath=checkpoint_path, save_weights_only=True, verbose=1)

    modelSummary = []

    # function api
    # 1) define input node / output node
    design = architecture.design
    inputs = keras.Input(shape=len(X_train[0]))
    print(inputs)
    print(len(X_train[0]))
    #input("arch")
    ran_initializer_output = random.choice(design.initializer)
    outputlayer = layers.Dense(4)

    # function for dense
    def funden(dense, x):
        return dense(x)

    # 2) create new node in layer for ANN network

    if architecture.layer == 1:
        ran_initializer = random.choice(design.initializer)
        ran_activation = random.choice(design.activations)
        # first hidden node
        dense = layers.Dense(architecture.units[0], kernel_initializer=ran_initializer, activation=ran_activation,
                             kernel_regularizer=design.regularizer)
        modelSummary.append(architecture.units[0])
        modelSummary.append(ran_initializer)
        modelSummary.append(ran_activation)
        finaloutput = funden(outputlayer, funden(dense, inputs))

    if architecture.layer > 1:
        denselist = []
        outputlist = []
        finaloutputs = [inputs]

        # Adding the input layer
        ran_initializer = random.choice(design.initializer)
        ran_activation = random.choice(design.activations)

        dense = layers.Dense(architecture.units[0], kernel_initializer=ran_initializer, activation=ran_activation,
                             kernel_regularizer=design.regularizer)
        outputlist.append(funden(dense, inputs))

        modelSummary.append(architecture.units[0])
        modelSummary.append(ran_initializer)
        modelSummary.append(ran_activation)

        for i in range(architecture.layer - 1):
            ran_initializer = random.choice(design.initializer)
            ran_activation = random.choice(design.activations)
            denselist.append(
                layers.Dense(architecture.units[i + 1], kernel_initializer=ran_initializer, activation=ran_activation,
                             kernel_regularizer=design.regularizer))
            modelSummary.append(architecture.units[i + 1])
            modelSummary.append(ran_initializer)
            modelSummary.append(ran_activation)
        print(len(denselist))
        for x in range(len(denselist)):
            finaloutputs.append(funden(denselist[x], finaloutputs[x]))
            output = finaloutputs[x + 1]
            print(output, x)

        # Adding the output layer
        finaloutput = funden(outputlayer, output)
        modelSummary.append("Output(4)")
        modelSummary.append(ran_initializer_output)

    model = keras.Model(inputs=inputs, outputs=finaloutput, name=cnnname)

    # Compiling the ANN
    ran_lossfunc = random.choice(design.lossfunctions)
    modelSummary.append(ran_lossfunc)
    model.compile(optimizer=design.optimizer,
                  loss=ran_lossfunc, metrics=['mse', 'accuracy'])

    print("<<<<<<CNN_space : %s>>>>>>>>>>>" % cnnname)
    model.summary()
    #keras.utils.plot_model(model,'modelsummary.png',show_shapes=True)

    early_stop = tf.keras.callbacks.EarlyStopping(monitor='val_loss', mode='min', patience=10)
    # Fitting the ANN to the Training set
    # regressor.fit(X_train, y_train, batch_size = 10, verbose=1, epochs = 1000, validation_split=0.2)
    start_time = datetime.datetime.now()

    print(X_train[0])
    print(X_train[0].shape)


    model.fit(X_train, y_train, batch_size=32, verbose=1, epochs=1000, validation_split=0.15, callbacks=[early_stop])
    end_time = datetime.datetime.now()
    train_time = (end_time - start_time).total_seconds()
    string_time = cnnname + " : " + str(train_time) + "\n"

    print(start_time, end_time, train_time, string_time)

    file = open("dataset_new_avg/elixir-modelsave/train_time.txt", "a")
    file.write(string_time)
    file.close()

    # save model
    savemodel_name = "dataset_new_avg/elixir-modelsave/%s_%s_%s.h5" % (resultFileName, outputType, cnnname)
    model.save(savemodel_name)

    ############ Predict & Test the Model ############
    real_output = np.array(X_test)
    inputs = real_output
    predicted_output = model.predict(inputs)
    outputs = np.array(y_test)
    print("CNN predict output")
    print(predicted_output)

    valid_inference = model.predict(valid_input)
    print(valid_inference)

    valid_output = np.array(valid_output)
    rmse_1 = np.square(valid_inference-valid_output)
    print(rmse_1)

    validsetlen = len(valid_output)
    rmse_annset = []
    for x in range(0,4):
        rmse_1_sum=0
        for y in range(0,validsetlen):
            rmse_1_sum=rmse_1_sum+rmse_1[y][x]
        rmse_annset.append(math.sqrt(rmse_1_sum)/validsetlen)
    print(rmse_annset)

    del model
    # print accuracy
    result = print_accuracy_all(predicted_output, y_test, "CNN")
    return result, modelSummary, rmse_annset



def sequentialTrain(X, Y, X_train, X_test, Y_train, Y_test, labels_name,outputType,result_name,
                    resultFileName,valid_input,valid_output):

    #test_inputfile = np.array(test_input,np.int32)
    #valid_input=np.array(valid_input,np.int32)
    #inf_filename = inf_filename

    #layers : Max later size (1~ layers)
    #nodes : Max node size (4~nodes)
    models = []
    results = []
    validset = []
    rmse_set = []
    # to here add lr, svr, rf --> results.add(lr, svr, rf)
    # 1. linearRegression
    # models.append(["LR",0,0,[" "]])  #put model`s architecutre like: get_arch()
    # result_1 = lrModel(X_train,X_test,Y_train,Y_test)
    # #result_1 = sklrModel(X_train, X_test, Y_train, Y_test)
    # results.append(result_1)
    # input("d")
    models.append(["LR_sklearn", 0, 0, [" "],[" "]])
    result, inference = linearRegression_sk(X_train, X_test, Y_train, Y_test,outputType,valid_input,valid_output)
    results.append(result)
    validset.append(inference)


    # print(result_1)
    # 2. svr
    models.append(["SVR",0,0,[" "],[" "]])  #put model`s architecutre like: get_arch()
    result, inference = lsvrModel(X_train, X_test, Y_train, Y_test,outputType,valid_input,valid_output)
    results.append(result)
    validset.append(inference)
    # 3. Random Forest
    rf_name = "RF_"
    # put model`s architecutre like: get_arch()
    for t in range(0,1):
        randomValue = random.randint(2,10)
        tree = 10
        nameV = rf_name+str(t+1)
        models.append([nameV,0,0,[tree],["TREE number: ",tree]])  #put model`s architecutre like: get_arch()
        result, inference = rforestModel(X_train, X_test, Y_train, Y_test,outputType,valid_input,valid_output,tree)
        results.append(result)
        validset.append(inference)
    print(inference)
    predict_col_1 = np.array(inference).T[0]
    predict_col_2 = np.array(inference).T[1]
    predict_col_3 = np.array(inference).T[2]
    predict_col_4 = np.array(inference).T[3]
    print(predict_col_1,predict_col_2,predict_col_3,predict_col_4)

    #inferencefile.save("dataset_new_avg/result/" + inf_filename)


    #0. EXTRA emsemble model : lightBGM, xGboost
    #1. these models are worked with single input data
    #2. need to split 1 input --> 4 inputs
    outputset = Y
    singleOutputset = []
    testsingleOutputset = []
    light_results = []
    xgboost_results = []
    lables = labels_name
    light_in = []
    xg_in = []
    for p in range(len(lables)):
        singleOutputset.append(outputset[lables[p]])
    print(singleOutputset)


    for e in range(len(lables)):
        e_train, e_test = train_test_split(singleOutputset[e], test_size=72, random_state=0)
        result, inference = lightgbm_model(X_train, X_test, e_train, e_test, valid_input, valid_output)
        light_results.append(result)
        light_in.append(inference)
        result, inference = XGBoost_model(X_train, X_test, e_train, e_test, valid_input, valid_output)
        xgboost_results.append(result)
        xg_in.append(inference)

    #wirteInferenceSingle(test_input, light_in, inferencefile, outputType, "LIGTGBM")
    #wirteInferenceSingle(test_input, xg_in, inferencefile, outputType, "XGBoost")
    models.append(["LightGBM",0,0,[" "],[" "]])
    results.append(light_results)
    light_in = np.transpose(light_in)
    validset.append(light_in)
    models.append(["XGBoost", 0, 0, [" "], [" "]])
    results.append(xgboost_results)
    xg_in = np.transpose(xg_in)
    validset.append(xg_in)
    #inferencefile.save("dataset_new_avg/result/" + inf_filename)


    print(models)
    print(results)
    print("iii")
    print(validset)
    valid_output=np.array(valid_output)
    print("len ")
    testlen = len(valid_output)
    print(testlen)
    print(valid_output)
    print("iii")


    real_val = []
    for mo in range(0,5): #model
        real_val.append(np.square(validset[mo]-valid_output))
    print(real_val)

    rmse_2_set=[]
    for aa in range(0,5):
        rmse_1_set=[]
        for ba in range(0,4):
            rmse_1=0
            for ca in range(0,testlen):
                rmse_1 = rmse_1+real_val[aa][ca][ba]
                print(aa,ba,ca,real_val[aa][ca][ba],rmse_1)
            rmse_1=math.sqrt(rmse_1)/testlen
            rmse_1_set.append(rmse_1)
        rmse_2_set.append(rmse_1_set)
    print(rmse_2_set)

    minrmse=[]
    rmsemodel=[]
    for tuo in range(0,4): #output
        minv=999999
        modelnum=0
        for model in range(0,5): #model
            if minv > rmse_2_set[model][tuo]:
                minv=rmse_2_set[model][tuo]
                modelnum=model
        minrmse.append(minv)
        rmsemodel.append(modelnum)
    print(minrmse)
    print(rmsemodel)

    for id in range(0,4):
        if rmsemodel[id]==0:
            rmsemodel[id]='LR'
        elif rmsemodel[id]==1:
            rmsemodel[id]='SVR'
        elif rmsemodel[id]==2:
            rmsemodel[id]='RF'
        elif rmsemodel[id]==3:
            rmsemodel[id]='XGBOOST'
        elif rmsemodel[id]==4:
            rmsemodel[id]='LIGHTGBM'
    print(rmsemodel)

    # #find min RMSE form LR, SVR, RF (shallow search)
    # for q in range(0,4):
    #     setall = []
    #     for w in range(0,5):
    #         setall.append(results[w][q][4])
    #     rmse_set.append(min(setall))
    # print("rmse_minset: ", str(rmse_set))


    # 4. ANN (deep search1)
    input_num = X.shape[1]
    output_num = Y.shape[1]

    name_count = 1
    architecture = Architecture("ANN", 0, 0, Default())
    startNode = [ [5, 6, 7, 4, 7, 6, 5, 4], [7, 4, 6, 5, 4, 7, 6, 5], [6, 5, 7, 4, 5, 6, 4, 7] ,[5, 6, 7, 4, 7, 6, 5, 4], [7, 4, 6, 5, 4, 7, 6, 5], [6, 5, 7, 4, 5, 6, 4, 7] ,[5, 6, 7, 4, 7, 6, 5, 4], [7, 4, 6, 5, 4, 7, 6, 5], [6, 5, 7, 4, 5, 6, 4, 7] ]
    print(startNode[0], startNode[1])
    for i in range(0,3): #20 --> 30
        cnn_randomnode(startNode[i], architecture, 4, models, results, X_train, X_test, Y_train, Y_test,
                       name_count, minrmse, rmsemodel, outputType,
                       resultFileName, valid_input, valid_output)

        earlystop_2 = ann_randomnode(startNode[i], architecture,4, models, results, X_train, X_test, Y_train, Y_test,name_count,minrmse,rmsemodel,outputType,
                                     resultFileName,valid_input,valid_output)
        name_count = name_count + 1
        if earlystop_2 == 1:
            break

    #inferencefile.remove(inferencefile['Sheet'])
    #inferencefile.save("dataset_new_avg/result/" + inf_filename)

    return models, results, rmsemodel, minrmse

def ann_randomnode(startNode,architecture,nodes,models,results,X_train, X_test, Y_train, Y_test,name_count,minrmse,rmsemodel,outputType,resultFileName,valid_input,valid_output):
    newNode=8
    index_node = 0
    for i in range(1, 11):  #range from 1 to max layer size (Layer size)
        architecture.initial_utnis([])
        annname = "ANN_" + str(name_count) + "_" + str(i)
        architecture.update_name(annname)
        newlayer = random.randint(1, 10)
        architecture.update_layers(newlayer)
        count = 1
        for j in range(0, newlayer):
            print("count: ",count)
            print("nodes(i): ",nodes)
            if newNode > 8:
                ranNode=7
            else:
                if count == 1:
                    newNode = startNode[index_node]
                    if index_node == 7:
                        index_node = 0
                    else:
                        index_node = index_node + 1
                else:
                    ranNode = random.randint(4, 7)
                    print("ranNode: ", ranNode)
                    print("newNode: ", newNode)
                    while (newNode < (7 + nodes - ranNode)):
                        print("while: ", newNode)
                        if newNode==4:
                            ranNode=7
                            break;
                        else:
                            ranNode = random.randint(4, 7)
                            print("reNUMBER: ",ranNode)
                            print("nodes(i): ", nodes)


                    newNode = 7+nodes - ranNode
            if newNode ==8:
                newNode = 7
            count = count+1
            print("random_value for newNode : ",newNode)
            architecture.update_nodes(newNode)
        #initializer_index = random.randint(1,2)
        #activation_index = random.randint(1, 3)
        result, modelSummary, annrmse = annModel(architecture, X_train, X_test, Y_train, Y_test,annname,outputType,resultFileName,valid_input,valid_output)
        if result == None:
            continue
        #wirteInference(test_inputfile, inference, inferencefile, outputType, annname)

        #inferencefile.save('/Users/yeonhooy/ccmodel/test_modelspace/' +"(PRE)" +inf_filename)
        architecture.update_modelSumary(modelSummary)
        print("ANN arc - ", architecture.get_arch())
        models.append(architecture.get_arch())
        results.append(result)
        #print(result[0][4], result[1][4], result[2][4], result[3][4])
        print("extra model`s best rmse: ",minrmse)
        print("ann`s best rmse: ",annrmse)
        if annrmse[0] < minrmse[0] and annrmse[1] < minrmse[1] and annrmse[2] < minrmse[2] and annrmse[3] < minrmse[3]:
            print("best result!!")
            minrmse[0] = annrmse[0]
            rmsemodel[0] = annname
            minrmse[1] = annrmse[1]
            rmsemodel[1] = annname
            minrmse[2] = annrmse[2]
            rmsemodel[2] = annname
            minrmse[3] = annrmse[3]
            rmsemodel[3] = annname
            print(rmsemodel)
            return 1
        else:
            if annrmse[0] < minrmse[0]:
                minrmse[0]=annrmse[0]
                rmsemodel[0]=annname
            if annrmse[1] < minrmse[1]:
                minrmse[1]=annrmse[1]
                rmsemodel[1] = annname
            if annrmse[2] < minrmse[2]:
                minrmse[2]=annrmse[2]
                rmsemodel[2] = annname
            if annrmse[3] < minrmse[3]:
                minrmse[3]=annrmse[3]
                rmsemodel[3] = annname
            print(rmsemodel)
    return 2

def cnn_randomnode(startNode,architecture,nodes,models,results,X_train, X_test, Y_train, Y_test,name_count,minrmse,rmsemodel,outputType,resultFileName,valid_input,valid_output):

    #scale
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)

    # conv layer
    # filter size
    # filter Num
    # stride size
    # pooling layer (max or average)
    convparams = []  # [filtersize, filterNum, strideSize, pooling layer)
    print("x_train.shape")
    print(X_train.shape)
    print(X_train.shape[0])

    for i in range(1, 11):  # layer size (max : 11)
        X_train_conv = []
        X_test_conv = []
        X_train_flat = []
        X_test_flat = []
        for q in range(0, X_train.shape[0]):
            print(X_train[q].shape)
            print(q)
            X_train_conv.append(X_train[q].reshape(1, 7, 1))
            print(X_train[q].reshape(1, 7, 1))
        for k in range(0, X_test.shape[0]):
            X_test_conv.append(X_test[k].reshape(1, 7, 1))
        convSummary = []
        intKernel = 7
        nextKernel = 0
        intFilter = 1
        kernelShape = (1, intKernel, intFilter)
        filterNum = 1
        print(resultFileName)
        convsumPath = 'dataset_new_avg/elixir-modelsave/'+resultFileName + '_' + outputType+ '_cnn_convsummary_elixir.txt'
        print(convsumPath)
        #input("resultfilename")
        convsumtxt = open(convsumPath, 'a')
        os.system('chmod 777 ' + convsumPath)
        fTrue = True
        while(fTrue):
            filter = random.randint(2, 7)
            print("filter: ",filter, "intKerenl: ",intKernel)
            print(convSummary)
            if filter >= intKernel:
                flat = 'yes'
                #flatten
                print(type(X_train_conv[0]))
                print(type(flat))
                if str(type(X_train_conv[0])) == "<class 'numpy.ndarray'>":
                    for ia in range(0, X_train.shape[0]):
                        # print("!!")
                        # print("type: ",type(X_train_conv[i]))
                        # print("c: ",X_train_conv[i])
                        # print(intKernel,nextKernel,filterNum)
                        X_train_flat.append(X_train_conv[ia].flatten())
                    for k in range(0, X_test.shape[0]):
                        X_test_flat.append(X_test_conv[k].flatten())
                else:
                    for ib in range(0, X_train.shape[0]):
                        # print("!!")
                        # print("type: ",type(X_train_conv[i]))
                        # print("c: ",X_train_conv[i])
                        # print(intKernel,nextKernel,filterNum)
                        X_train_flat.append(X_train_conv[ib].numpy().flatten())
                    for k in range(0, X_test.shape[0]):
                        X_test_flat.append(X_test_conv[k].numpy().flatten())
                acf='flat'
                stride = 1
                convSummary.append((filterNum,filter,acf,stride,flat,intKernel,nextKernel))
                break
                convnode = intKernel * filterNum
                fTrue = False
            else:
                flat = 'no'
                convnode = 7
                fTrue = True
            Strue = True
            while(Strue):
                stride = random.randint(1,filter)
                quot, remain = divmod(intKernel-filter, stride)
                if remain == 0:
                    nextKernel = quot+1
                    Strue = False
                else:
                    Strue = True
                    print("Re search")


            filterNum = random.randint(2, 100)
            convAfunc = random.randint(1,2)
            if convAfunc == 1:
                    acf = 'elu'
            elif convAfunc == 2:
                    acf = 'relu'
            outShpae = (1,nextKernel,filterNum)

            convSummary.append((filterNum,filter,acf,stride,flat,intKernel,nextKernel))
            print(convSummary)
            # for j in range(0, X_train.shape[0]):
            #     X_train_conv[j] = Conv1D(filters=filterNum,kernel_size=filter,activation=acf,strides=stride,input_shape=kernelShape)(X_train_conv[j])
            # for m in range(0, X_test.shape[0]):
            #     X_test_conv[m] = Conv1D(filters=filterNum, kernel_size=filter, activation=acf, strides=stride, input_shape=kernelShape)(X_test_conv[m])
            print(X_train_conv[0])
            intKernel = nextKernel
            intFilter = filterNum
            kernelShape = outShpae
            print(X_train)
            print(X_train.shape)

        architecture.initial_utnis([])
        cnnname = "CNN_" + str(name_count) + "_" + str(i)

        #save convolution parameters
        convsumtxt.write(cnnname)
        convsumtxt.write('/')
        convsumtxt.write(str(len(convSummary)))
        convsumtxt.write('/')
        convsumtxt.write(str(convSummary))
        convsumtxt.write('/')
        convsumtxt.write('\n')
        convsumtxt.close()

        print(convSummary)
        #input("conv")

        X_train_flat, X_test_flat, X_valid_flat, convnode = convolutionLayer(X_train,X_test,valid_input,cnnname,resultFileName,outputType)

        newNode = convnode
        index_node = 0

        architecture.update_name(cnnname)

        newlayer = random.randint(1, 10)
        architecture.update_layers(newlayer)
        count = 1
        for j in range(0, newlayer):
            print("count: ", count)
            print("nodes(i): ", nodes)
            if newNode <= 0:
                ranNode = convnode
            else:
                if count == 1:
                    newNode = convnode
                    #newNode = startNode[index_node]
                    if index_node == newlayer:
                        index_node = 0
                    else:
                        index_node = index_node + 1
                else:
                    ranNode = random.randint(1, convnode)
                    print("ranNode: ", ranNode)
                    print("newNode: ", newNode)
                    while (newNode < (convnode + nodes - ranNode)):
                        print("while: ", newNode)
                        if newNode == 1:
                            ranNode = convnode
                            break;
                        else:
                            ranNode = random.randint(1, convnode)
                            print("reNUMBER: ", ranNode)
                            print("nodes(i): ", nodes)

                    newNode = convnode + nodes - ranNode
            if newNode < nodes:
                newNode = nodes
            count = count + 1
            print("random_value for newNode : ", newNode)
            #input("dds")
            architecture.update_nodes(newNode)
        # initializer_index = random.randint(1,2)
        # activation_index = random.randint(1, 3)
        print("hel")
        print(X_train.shape)
        print(type(X_train_flat))

        print(type(X_train))
        print(type(X_train_flat))

        #input("dsas")
        result, modelSummary, cnnrmse = cnnModel(architecture, X_train_flat, X_test_flat, Y_train, Y_test, cnnname, outputType,
                                                 resultFileName, X_valid_flat, valid_output)
        if result == None:
            continue
        # wirteInference(test_inputfile, inference, inferencefile, outputType, annname)

        # inferencefile.save('/Users/yeonhooy/ccmodel/test_modelspace/' +"(PRE)" +inf_filename)
        architecture.update_modelSumary(modelSummary)
        print("CNN arc - ", architecture.get_arch())
        models.append(architecture.get_arch())
        results.append(result)
        # print(result[0][4], result[1][4], result[2][4], result[3][4])
        print("extra model`s best rmse: ", minrmse)
        print("cnn`s best rmse: ", cnnrmse)
        if cnnrmse[0] < minrmse[0] and cnnrmse[1] < minrmse[1] and cnnrmse[2] < minrmse[2] and cnnrmse[3] < minrmse[3]:
            print("best result!!")
            minrmse[0] = cnnrmse[0]
            rmsemodel[0] = cnnname
            minrmse[1] = cnnrmse[1]
            rmsemodel[1] = cnnname
            minrmse[2] = cnnrmse[2]
            rmsemodel[2] = cnnname
            minrmse[3] = cnnrmse[3]
            rmsemodel[3] = cnnname
            print(rmsemodel)
            return 1
        else:
            if cnnrmse[0] < minrmse[0]:
                minrmse[0] = cnnrmse[0]
                rmsemodel[0] = cnnname
            if cnnrmse[1] < minrmse[1]:
                minrmse[1] = cnnrmse[1]
                rmsemodel[1] = cnnname
            if cnnrmse[2] < minrmse[2]:
                minrmse[2] = cnnrmse[2]
                rmsemodel[2] = cnnname
            if cnnrmse[3] < minrmse[3]:
                minrmse[3] = cnnrmse[3]
                rmsemodel[3] = cnnname
            print(rmsemodel)
    return 2

def convolutionLayer(X_train,X_test,X_valid,cnnname,resultFileName,outputType):
    # scale
    scaler = MinMaxScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.fit_transform(X_test)
    X_valid = scaler.fit_transform(X_valid)
    X_train_conv = []
    X_test_conv = []
    X_valid_conv = []
    X_train_flat = []
    X_test_flat = []
    X_valid_flat = []
    # conv layer
    # filter size
    # filter Num
    # stride size
    # pooling layer (max or average)
    convparams = []  # [filtersize, filterNum, strideSize, pooling layer)
    print("x_train.shape")
    print(X_train.shape)
    print(X_train.shape[0])

    convSummary = []
    intKernel = 7
    nextKernel = 7
    intFilter = 1
    kernelShape = (1, intKernel, intFilter)
    filterNum = 1

    #1. data convert reshape: 3d form (1,7,1)
    for q in range(0, X_train.shape[0]):
        print(X_train[q].shape)
        print(q)
        X_train_conv.append(X_train[q].reshape(1, 7, 1))
        print(X_train[q].reshape(1, 7, 1))
    for k in range(0, X_test.shape[0]):
        X_test_conv.append(X_test[k].reshape(1, 7, 1))
    for v in range(0, X_valid.shape[0]):
        X_valid_conv.append(X_valid[v].reshape(1, 7, 1))

    #2. read conv summary

    while(True):
        convsumPath = 'dataset_new_avg/elixir-modelsave/'+resultFileName + '_' + outputType+ '_cnn_convsummary_elixir.txt'
        convsumtxt = open(convsumPath,'r')
        sumlines = []
        thisconvSum = []
        findname = 0
        for paragraph in convsumtxt:
            sumlines = paragraph.split("/")
            for each_line in sumlines:
                if each_line==cnnname:
                    thisconvSum=sumlines
                    findname = findname+1
                    break
                else:
                    pass
        convsumtxt.close()
        print(cnnname)
        if findname > 0:
            break
    print(thisconvSum)
    layersummary = thisconvSum[2]
    cx = "[]()"
    arraysummary = ''.join(x for x in layersummary if x not in cx)
    print(arraysummary)
    summaries = arraysummary.split(", ")
    print(summaries)
    print(len(summaries))
    layertupples = []
    tuplen = int(len(summaries)/7)
    for q in range(tuplen):
        tup = (summaries[0+7*q],summaries[1+7*q],summaries[2+7*q],summaries[3+7*q],summaries[4+7*q],summaries[5+7*q],summaries[6+7*q])
        layertupples.append(tup)
    print(layertupples)

    kernelShape = (1,intKernel, intFilter)
    if tuplen >1:
        for tu in range(tuplen-1):
            filterNum = int(layertupples[tu][0])
            filter = int(layertupples[tu][1])
            acf = str(layertupples[tu][2])
            stride = int(layertupples[tu][3])
            flatcheck = str(layertupples[tu][4])
            intKernel = int(layertupples[tu][5])
            nextKernel = int(layertupples[tu][6])
            outShpae = (1, nextKernel, filterNum)
            print(acf)
            print(type(acf))

            for j in range(0, X_train.shape[0]):
                X_train_conv[j] = Conv1D(filters=filterNum,kernel_size=filter,strides=stride,input_shape=kernelShape)(X_train_conv[j])
            for m in range(0, X_test.shape[0]):
                X_test_conv[m] = Conv1D(filters=filterNum, kernel_size=filter, strides=stride, input_shape=kernelShape)(X_test_conv[m])
            for v in range(0, X_valid.shape[0]):
                X_valid_conv[v] = Conv1D(filters=filterNum, kernel_size=filter, strides=stride, input_shape=kernelShape)(X_valid_conv[v])
            print(X_train_conv[0])
            kernelShape = outShpae

    lastNodes = nextKernel * filterNum
    if str(type(X_train_conv[0])) == "<class 'numpy.ndarray'>":
        for ia in range(0, X_train.shape[0]):
            X_train_flat.append(X_train_conv[ia].flatten())
            #X_train_flat.append(X_train_conv[ia].reshape(1,lastNodes))
        for k in range(0, X_test.shape[0]):
            X_test_flat.append(X_test_conv[k].flatten())
            #X_test_flat.append(X_test_conv[k].reshape(1,lastNodes))
        for v in range(0, X_valid.shape[0]):
            X_valid_flat.append(X_valid_conv[v].flatten())
    else:
        for ib in range(0, X_train.shape[0]):
            X_train_flat.append(X_train_conv[ib].numpy().flatten())
            #X_train_flat.append(X_train_conv[ib].numpy().reshape(1,lastNodes))
        for k in range(0, X_test.shape[0]):
            X_test_flat.append(X_test_conv[k].numpy().flatten())
            #X_test_flat.append(X_test_conv[k].numpy().reshape(1,lastNodes))
        for v in range(0, X_valid.shape[0]):
            X_valid_flat.append(X_valid_conv[v].numpy().flatten())


    print(X_train_flat[0])
    print(len(X_train_flat[0]))
    print(len(X_train_flat))
    print(layertupples)
    convnode = len(X_train_flat[0])
    print(convnode,lastNodes)
    #input("check")
    print(type(X_train_flat[0]))
    print(X_train_flat[0].shape)
    print(convnode)
    #input("broken")

    X_train_flat = np.array(X_train_flat)
    X_test_flat = np.array(X_test_flat)
    X_valid_flat = np.array(X_valid_flat)

    return X_train_flat, X_test_flat, X_valid_flat, convnode

def wirteInference(testdatafile,inference,resultfile,outputType,modelName):
    testdata = np.array(testdatafile)
    print(testdata)


    if outputType == 'TX':
        # 모델 당 sheet 생성
        sheet = resultfile.create_sheet(modelName)
        print(resultfile.get_sheet_names())
        # input("DD")
        sheet.cell(row=1, column=1).value = "Switch_e"
        sheet.cell(row=1, column=2).value = "Switch_c"
        sheet.cell(row=1, column=3).value = "host"
        sheet.cell(row=1, column=4).value = "connection"
        sheet.cell(row=1, column=5).value = "interval"
        sheet.cell(row=1, column=6).value = "link"
        sheet.cell(row=1, column=7).value = "hop"
        for t in range(0, len(testdata)):
            for p in range(0, 7):
                sheet.cell(row=2 + t, column=1 + p).value = testdata[t][p]
        sheet.cell(row=1, column=8).value = "avgSentMsg"
        sheet.cell(row=1, column=9).value = "avgSentByte"
        sheet.cell(row=1, column=10).value = "secMaxSendMsg"
        sheet.cell(row=1, column=11).value = "secMaxSendByte"
        for i in range(0,len(inference)):
            for k in range(0,4):
                sheet.cell(row=2+i, column=8+k).value = inference[i][k]

    if outputType == 'RX':
        print(resultfile.get_sheet_names())
        # input("DD")
        try:
            sheet = resultfile.get_sheet_by_name(modelName)
        except:
            sheet = resultfile.create_sheet(modelName)
            sheet.cell(row=1, column=1).value = "Switch_e"
            sheet.cell(row=1, column=2).value = "Switch_c"
            sheet.cell(row=1, column=3).value = "host"
            sheet.cell(row=1, column=4).value = "connection"
            sheet.cell(row=1, column=5).value = "interval"
            sheet.cell(row=1, column=6).value = "link"
            sheet.cell(row=1, column=7).value = "hop"
            for t in range(0, len(testdata)):
                for p in range(0, 7):
                    sheet.cell(row=2 + t, column=1 + p).value = testdata[t][p]
        sheet.cell(row=1, column=12).value = "avgRecvMsg"
        sheet.cell(row=1, column=13).value = "avgRecvByte"
        sheet.cell(row=1, column=14).value = "secMaxRecvMsg"
        sheet.cell(row=1, column=15).value = "secMaxRecvByte"
        for j in range(0, len(inference)):
            for l in range(0, 4):
                sheet.cell(row=2 + j, column=12 + l).value = inference[j][l]


def wirteInferenceSingle(testdatafile,inference,resultfile,outputType,modelName):
    testdata = np.array(testdatafile)
    print(testdata)

    if outputType == 'TX':
        # create sheet per model
        sheet = resultfile.create_sheet(modelName)
        sheet.cell(row=1, column=1).value = "Switch_e"
        sheet.cell(row=1, column=2).value = "Switch_c"
        sheet.cell(row=1, column=3).value = "host"
        sheet.cell(row=1, column=4).value = "connection"
        sheet.cell(row=1, column=5).value = "interval"
        sheet.cell(row=1, column=6).value = "link"
        sheet.cell(row=1, column=7).value = "hop"
        for t in range(0, len(testdata)):
            for p in range(0, 7):
                sheet.cell(row=2 + t, column=1 + p).value = testdata[t][p]

        sheet.cell(row=1, column=8).value = "avgSentMsg"
        sheet.cell(row=1, column=9).value = "avgSentByte"
        sheet.cell(row=1, column=10).value = "secMaxSendMsg"
        sheet.cell(row=1, column=11).value = "secMaxSendByte"
        for i in range(0,len(inference[0])):
            for k in range(0,4):
                sheet.cell(row=2+i, column=8+k).value = inference[k][i]

    if outputType == 'RX':
        sheet = resultfile[modelName]
        sheet.cell(row=1, column=12).value = "avgRecvMsg"
        sheet.cell(row=1, column=13).value = "avgRecvByte"
        sheet.cell(row=1, column=14).value = "secMaxRecvMsg"
        sheet.cell(row=1, column=15).value = "secMaxRecvByte"
        for j in range(0, len(inference[0])):
            for l in range(0, 4):
                sheet.cell(row=2 + j, column=12 + l).value = inference[l][j]


def csvTrain_withoutKfold(result_excel, X,Y, result_name,worksheetName,outputType,resultFileName,scoring_input,scoring_output):

    print("CSVTRAIN",X,Y)
    all_results = []
    X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=72, random_state=0)
    print(X_train, X_test, Y_train, Y_test)
    scoring_input = pd.concat([X_test,scoring_input])
    scoring_output = pd.concat([Y_test,scoring_output])
    print(scoring_input.head(),scoring_output.head())
    #input("test-scoring")
    models, results, rmsemodel, minrmse = sequentialTrain(X, Y, X_train, X_test, Y_train, Y_test,worksheetName,outputType,result_name,
                                      resultFileName,scoring_input,scoring_output)  # 3,3 --> (1,2,4) , (1,2,4)
    print("MODELS",models)
    print("RESULTS",results)
    print("BESTMDOELS",rmsemodel)
    write_excel(result_excel, models, results, result_name, worksheetName,rmsemodel,minrmse)



def csvTrain(result_excel, X,Y, result_name,worksheetName):


    print("CSVTRAIN",X,Y)
    all_results = []
    # 1-3. Split Test / Train Data
    n = 5
    kf = KFold(n_splits=n)
    count = 1
    for train_index, test_index in kf.split(X,Y):
        #print("TRAIN:", train_index, "TEST:", test_index)
        X_train, X_test = X.iloc[train_index], X.iloc[test_index]
        Y_train, Y_test = Y.iloc[train_index], Y.iloc[test_index]
        #X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.2, random_state=0)
        print(X_train, X_test, Y_train, Y_test)

        models, results = sequentialTrain(X, Y, X_train, X_test, Y_train, Y_test)  # 3,3 --> (1,2,4) , (1,2,4)
        print("COUNT: ",count)
        all_results.append(results)
        count = count+1
    model_length = len(models)
    total_elements = model_length*36
    tmp = np.arange(total_elements).reshape(model_length,4,9)
    #initial tmp[]
    for nx in range(0,model_length):
        for ny in range(0,4):
            for nz in range(0,9):
                tmp[nx][ny][nz] = 0
    tmp_n = np.arange(total_elements).reshape(model_length, 4, 9)
    # initial tmp[]
    for nx in range(0, model_length):
        for ny in range(0, 4):
            for nz in range(0, 9):
                tmp_n[nx][ny][nz] = 0
    print("init_TMP : ",tmp)
    print("LENGTH OF ALL_RESULTS : ", len(all_results))
    for i in range(len(all_results)):
        #tmp = tmp + all_results[i]
        print("Added all_results[i]", all_results[i])
        result_tmp = all_results[i]
        for a in range(model_length):
            for b in range(0,4):
                for c in range(0, 9):
                    if result_tmp[a][b][c] != np.inf :
                        print(result_tmp[a][b][c], type(result_tmp[a][b][c]))
                        tmp[a][b][c] = tmp[a][b][c] + result_tmp[a][b][c]
                    elif result_tmp[a][b][c] == np.nan:
                        print("NAN", result_tmp[a][b][c])
                    else:
                        print("INF",result_tmp[a][b][c])

    print("tmp", tmp)
    tmp_n = np.round(tmp/n,2)
    print("tmp_n",tmp_n)
    print("Modles: ",models)

    write_excel(result_excel, models, tmp_n, result_name,worksheetName)


def main():
                                                ##########################################
                                                ################ MAIN ####################
                                                ##########################################
    # cuda setting
    os.environ["CUDA_VISIBLE_DEVICES"] = "0,1,2,3"
    # file_name = input('file name. >> ')
    # data_name = file_name + '.csv'
    # result_name = file_name + '.xlsx'
    pd.options.display.float_format = '{:.5f}'.format

    from tensorflow.python.client import device_lib

    print(device_lib.list_local_devices())
    mirrored_strategy = tf.distribute.MirroredStrategy(devices=["/gpu:0", "/gpu:1", "/gpu:2", "/gpu:3"])
    print('Number of devices: {}'.format(mirrored_strategy.num_replicas_in_sync))
    print("GPU is available? ",tf.test.is_gpu_available())
    print(tf.config.list_physical_devices('gpu'))

    #0. get csv files in the folder.
    path_name = "dataset_new_avg"
    file_list = os.listdir(path_name)
    file_list_csv = [file for file in file_list if file.endswith(".csv")]
    file_list_csv.sort()
    print(file_list_csv)
    counter = 1

    # 0r, input training set in argv
    training_datset = sys.argv[1:]
    print(training_datset)
    file_list_csv=training_datset




    #1. get test data (for inference) - will not be used at training session
    test_path_name = "dataset_new_avg/test_datacenter"
    testfile_list = os.listdir(test_path_name)
    testfile_list_csv = [file for file in testfile_list if file.endswith(".csv")]
    testfile_list_csv.sort()
    print(testfile_list_csv)


    for csvfile in file_list_csv:
        filenaming = csvfile.split('.')
        resultFileName = filenaming[0]
        print("Result file name : "+resultFileName)
        print("Now train: {}", path_name + "/" + csvfile)


        # 1. Read Data and remove outlier data
        print("++++++ Read Data ++++++")
        print(csvfile)
        dataset = pd.read_csv('dataset_new_avg/'+csvfile)
        dataset = dataset.dropna()
        print(dataset.head())
        print("origin dataset length",len(dataset))


        X = dataset.drop(
            labels=['time', 'avgSentMsg', 'avgSentByte', 'secMaxSendMsg', 'secMaxSendByte', 'avgRecvMsg', 'avgRecvByte',
                    'secMaxRecvMsg', 'secMaxRecvByte'], axis=1)  # 7개의 Input data / label
        # output for TX / RX
        y_1 = dataset[['avgSentMsg', 'avgSentByte', 'secMaxSendMsg', 'secMaxSendByte']]
        y_2 = dataset[['avgRecvMsg', 'avgRecvByte', 'secMaxRecvMsg', 'secMaxRecvByte']]

        # output for msg / byte
        # y_3 = dataset[['avgSentMsg', 'secMaxSendMsg', 'avgRecvMsg', 'secMaxRecvMsg']]
        # y_4 = dataset[['avgSentByte', 'secMaxSendByte', 'avgRecvByte', 'secMaxRecvByte']]

        #y = [y_1, y_2, y_3, y_4]
        y = [y_1, y_2]

        print(X)
        print(y_1)
        print(y_2)
        # print(y_3)
        # print(y_4)

        #validset for supreme model
        realvalidset = pd.read_csv('dataset_new_avg/scoring/'+csvfile)
        scoring_input = realvalidset[['switch_e','switch_c','host','connection','interval','link','hop']]
        scoring_output = [realvalidset[['avgSentMsg', 'avgSentByte', 'secMaxSendMsg', 'secMaxSendByte']],
                        realvalidset[['avgRecvMsg', 'avgRecvByte', 'secMaxRecvMsg', 'secMaxRecvByte']]]


        for testfile in testfile_list_csv:
            testfilenaming = testfile.split('.')
            testtFileName = testfilenaming[0]
            pathfullName = 'dataset_new_avg/test_datacenter/'+testtFileName + '.csv'
            print("test file name : " + pathfullName)

            testdataset = pd.read_csv(pathfullName)
            print(testdataset)

        test_input = testdataset[['switch_e','switch_c','host','connection','interval','link','hop']]




        resultNum = ["TX","RX","MSG","Byte"]
        resultfilename = ["","","",""]
        outputTitles = [['avgSentMsg', 'avgSentByte', 'secMaxSendMsg', 'secMaxSendByte'],['avgRecvMsg', 'avgRecvByte', 'secMaxRecvMsg', 'secMaxRecvByte'],
                        ['avgSentMsg', 'secMaxSendMsg', 'avgRecvMsg', 'secMaxRecvMsg'],['avgSentByte', 'secMaxSendByte', 'avgRecvByte', 'secMaxRecvByte']]
        print("OutputitleTest",outputTitles[0])

        # create inference file
        import openpyxl
        inferencefile = openpyxl.Workbook()
        test_inputfile = np.array(test_input, np.int32)
        inf_filename = resultFileName + "_inference.xlsx"
        inferencefile.save("dataset_new_avg/result/datcenter/" + inf_filename)

        #0. create timestamp file
        file = open("dataset_new_avg/timestamp/"+resultFileName+".txt", "w")
        file.close()

        # 1 round -> tx, 2 round -> rx
        for num in range(0,2):
            result_excel = Workbook()
            resultfilename[num] = resultFileName+"_result_"+resultNum[num]+".xlsx"
            print("resultNAME: ",resultfilename[num])
            # csvTrain(result_excel, X, y[num], resultfilename[num],outputTitles[num])

            start_time = datetime.datetime.now()
            csvTrain_withoutKfold(result_excel, X, y[num], resultfilename[num],outputTitles[num],resultNum[num],resultFileName,
                                  scoring_input,scoring_output[num])


            end_time = datetime.datetime.now()
            train_time = (end_time - start_time).total_seconds()
            string_time = resultNum[num] + " : " + str(train_time) + "\n"

            print(start_time, end_time, train_time, string_time)

            file = open("dataset_new_avg/timestamp/"+resultFileName+".txt", "a")
            file.write(string_time)
            file.close()

            # result_excel.remove(result_excel['Sheet'])
            result_excel.save("dataset_new_avg/result/datcenter/" + resultfilename[num])
            #inferencefile.remove(inferencefile['Sheet'])
            inferencefile.save("dataset_new_avg/result/datcenter/" + inf_filename)
            counter = counter + 1
            print("COUNTER: ", counter)
    print("Finished!")

main()








