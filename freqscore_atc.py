from collections import Counter

import openpyxl
from openpyxl.styles import PatternFill, Color
import os


def supermeModelselction(xfile,modelSavefile,filename,modelDic):
    wb = openpyxl.load_workbook(xfile)
    sheetnames = wb.sheetnames
    outputcol = 0
    modelSavefile.cell(row=2,column=1).value = 'selectModel'
    modelSavefile.cell(row=3, column=1).value = 'freqRatio'

    for output in sheetnames:
        # selectModel=''
        # fileoutputName = ''
        if output == "Sheet":
            continue
        else:
            fileoutputName = filename + "+" + output
            selectModel = modelDic[fileoutputName]
        print(fileoutputName)
        print(selectModel)
        outputcol = outputcol+1
        modelSavefile.cell(row=1, column=1+outputcol).value = output
        sheet = wb[output]
        maxRow=sheet.max_row
        maxCol=sheet.max_column
        if maxCol == 1:
            pass
        else:
            print(maxRow, maxCol)
            sampleNum = maxRow - 1
            modelNum = maxCol - 8

            freqModels = robustfrequency(sheet,sampleNum,modelNum)
            freqModels,count=frequency_sort(freqModels)
            selectFreqRatio = frequency_selected(freqModels,selectModel)
            print(freqModels, count)
            print(selectFreqRatio)
            freqratio = round(100*(count/sampleNum),2)
            print(freqModels[0],freqratio)
            roFreqModel = freqModels[0]
            roRmseModel, minRmse, rmseDict= robustRMSE(sheet,sampleNum,modelNum)
            if freqratio > 44:
                supremeModel = roFreqModel
            else:
                supremeModel = roRmseModel
            print(supremeModel)
            #input("fr")
            modelSavefile.cell(row=2,column=1+outputcol).value=selectModel
            modelSavefile.cell(row=3, column=1 + outputcol).value = selectFreqRatio






def robustfrequency(sheet,sampleNum,modelNum):
    freqModels = []
    for j in range(sampleNum):
        realvalue = sheet.cell(row=2 + j, column=8).value
        errors = []
        for i in range(modelNum):
            predict = sheet.cell(row=2 + j,column=9+i).value
            print(i,j,realvalue,predict)
            error = abs(int(realvalue)-int(predict))
            errors.append(error)
        modelIndex = errors.index(min(errors))
        modelName = sheet.cell(row=1,column=9+modelIndex).value
        freqModels.append(modelName)
    return freqModels

def frequency_sort(data):
    rt_data=[]
    count = 0
    for d,c in Counter(data).most_common():
        for i in range(c):
            rt_data.append(d)
    print(rt_data)
    freqModel = rt_data[0]
    for i in rt_data:
        if i == freqModel:
            count = count+1
    return rt_data, count

def frequency_selected(data, selectModel):
    rt_data=[]
    total_len = len(data)
    count = 0
    for d,c in Counter(data).most_common():
        for i in range(c):
            rt_data.append(d)
    print(rt_data)
    freqModel = selectModel
    for i in rt_data:
        if i == freqModel:
            count = count+1
    ratio = count/float(total_len)*100
    return ratio

def robustRMSE(sheet,sampleNum,modelNum):
    rmseModels = []
    rmselist = []
    rmsedict = {}
    for i in range(modelNum):
        sum=0
        for j in range(sampleNum):
            realvalue = sheet.cell(row=2+j,column=8).value
            predict = sheet.cell(row=2+j,column=9+i).value
            m = (realvalue-predict)*(realvalue-predict)
            sum = sum+m
        rmsevalue = (sum/sampleNum)**(1/2)
        rmselist.append(rmsevalue)
        modelLabel = sheet.cell(row=1,column=9+i).value
        rmsedict[modelLabel] = rmsevalue
    modelIndex = rmselist.index(min(rmselist))
    modelName = sheet.cell(row=1, column=9 + modelIndex).value
    return modelName, min(rmselist), rmsedict

def loadSupreme(modelpath):
    modeltxt = open(modelpath,'r')
    models = {}
    while True:
        line = modeltxt.readline()
        if not line: break
        info = line.split(":")
        modelName = info[1].split("\n")
        models[info[0]] = modelName[0]
    return models

#MAIN
#파일 읽어오기
#path_name = "dataset_new_avg/result/evaluation_result"
#path_name = "dataset_new_avg/result/scoring_result"
path_name = "dataset_new_avg/result/naive_result"
file_list = os.listdir(path_name)
# resultfile이 TX인지 RX인지 구분
file_list = [file for file in file_list if file.endswith("inference.xlsx")]
file_list.sort()

outputs = ['avgSentMsg', 'avgSentByte', 'secMaxSendMsg','secMaxSendByte','avgRecvMsg','avgRecvByte','secMaxRecvMsg','secMaxRecvByte']

#결과 파일 생성
summarywb = openpyxl.Workbook()
modelDict = loadSupreme('dataset_new_avg/result/modelselect/'+'freq.txt') # supreme, naiive, freq

#한파일씩 처리하기(TX,RX 같이 처리하기)
for xlfile in file_list:
    #suptxt = open("dataset_new_avg/result/modelselect/supreme.txt", 'a')

    #pathfile = 'dataset_new_avg/result/evaluation_result/'+xlfile
    #pathfile = 'dataset_new_avg/result/scoring_result/' + xlfile
    pathfile = 'dataset_new_avg/result/naive_result/' + xlfile


    filenaming = xlfile.split('_inference.xlsx')
    resultFileName = filenaming[0]
    resultsheet = summarywb.create_sheet(resultFileName)
    supermeModelselction(pathfile,resultsheet,resultFileName,modelDict)
    #selectedModelResut(pathfile, resultsheet, resultFileName, modelDict)
    #summary_all(resultFileName,resultsheet)
    #summarywb.remove(summarywb['Sheet'])
    #suptxt.close()
summarywb.save('dataset_new_avg/result/modelselect/f_supremeSelectModel_freq(dt).xlsx')
print("FINISH!")









